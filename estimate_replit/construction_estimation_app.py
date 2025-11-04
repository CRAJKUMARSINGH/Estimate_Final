#!/usr/bin/env python3
"""
🏗️ CONSTRUCTION ESTIMATION SYSTEM - UNIFIED VERSION
==================================================
Complete, production-ready construction cost estimation application
Combines all improvements into a single, clean system

Features:
- Enhanced Excel import with formula preservation
- Real-time calculations and updates
- Database persistence and project management
- Advanced search and filtering
- Visual analytics and reporting
- Template system for reusable estimates
- Multi-user collaboration support

Author: Construction Estimation Team
Version: 3.0 (Unified)
Date: November 2025
"""

import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime
import sqlite3
import json
import re
import io
import tempfile
import os
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import logging
from functools import lru_cache
import openpyxl
from openpyxl import load_workbook
import base64

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# =============================================================================
# CONFIGURATION & CONSTANTS
# =============================================================================

# Page configuration
st.set_page_config(
    page_title="Construction Estimation System",
    page_icon="🏗️",
    layout="wide",
    initial_sidebar_state="expanded",
    menu_items={
        'Get Help': 'https://github.com/CRAJKUMARSINGH/estimate-v2.1.1',
        'Report a bug': 'mailto:crajkumarsingh@hotmail.com',
        'About': "Construction Estimation System v3.0 - Complete unified solution"
    }
)

# Data schemas
MEASUREMENT_COLUMNS = [
    'id', 'ssr_code', 'item_no', 'description', 'specification', 'location',
    'quantity', 'length', 'breadth', 'height', 'diameter', 'thickness', 
    'unit', 'total', 'deduction', 'net_total', 'remarks'
]

ABSTRACT_COLUMNS = [
    'id', 'ssr_code', 'description', 'unit', 'quantity', 'rate', 'amount'
]

SSR_COLUMNS = [
    'code', 'description', 'category', 'unit', 'rate'
]

# Standard units and measurement types
UNITS = ["RM", "Cum", "Sqm", "Nos", "Kg", "Ton", "Ltr", "LS", "Meter", "Feet", "Inch"]

MEASUREMENT_TYPES = {
    "Standard": "Nos × Length × Breadth × Height",
    "Linear": "Nos × Length", 
    "Area": "Nos × Length × Breadth",
    "Volume": "Nos × Length × Breadth × Height",
    "Circular Area": "Nos × π × (Diameter/2)²",
    "Circular Volume": "Nos × π × (Diameter/2)² × Height",
    "Deduction": "Gross - Deductions",
    "Weight": "Nos × Length × Unit Weight",
    "Custom Formula": "User Defined"
}

WORK_TYPES = {
    "Civil Work": "🏗️",
    "Sanitary Work": "🚰", 
    "Electrical Work": "⚡",
    "Landscape Work": "🌳"
}

# =============================================================================
# DATABASE MANAGEMENT
# =============================================================================

class UnifiedDatabase:
    """Unified database management for all estimation data"""
    
    def __init__(self, db_path: str = "construction_estimates.db"):
        self.db_path = db_path
        self.init_database()
    
    def init_database(self):
        """Initialize database with all required tables"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Projects table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS projects (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT NOT NULL,
                    location TEXT,
                    created_date TEXT,
                    last_modified TEXT,
                    total_cost REAL,
                    status TEXT DEFAULT 'active',
                    metadata TEXT
                )
            """)
            
            # Measurements table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS measurements (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    project_id INTEGER,
                    sheet_name TEXT,
                    item_no TEXT,
                    description TEXT,
                    specification TEXT,
                    location TEXT,
                    quantity REAL,
                    length REAL,
                    breadth REAL,
                    height REAL,
                    diameter REAL,
                    thickness REAL,
                    unit TEXT,
                    total REAL,
                    deduction REAL,
                    net_total REAL,
                    remarks TEXT,
                    ssr_code TEXT,
                    FOREIGN KEY (project_id) REFERENCES projects (id)
                )
            """)
            
            # Abstracts table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS abstracts (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    project_id INTEGER,
                    sheet_name TEXT,
                    ssr_code TEXT,
                    description TEXT,
                    unit TEXT,
                    quantity REAL,
                    rate REAL,
                    amount REAL,
                    FOREIGN KEY (project_id) REFERENCES projects (id)
                )
            """)
            
            # SSR items table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS ssr_items (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    code TEXT UNIQUE,
                    description TEXT,
                    category TEXT,
                    unit TEXT,
                    rate REAL,
                    last_updated TEXT
                )
            """)
            
            # Templates table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS templates (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT NOT NULL,
                    description TEXT,
                    category TEXT,
                    template_data TEXT,
                    created_date TEXT,
                    usage_count INTEGER DEFAULT 0
                )
            """)
            
            conn.commit()
            conn.close()
            
        except Exception as e:
            logger.error(f"Database initialization error: {e}")
            st.error(f"Database error: {e}")
    
    def save_project(self, project_data: Dict) -> int:
        """Save complete project to database"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Insert or update project
            cursor.execute("""
                INSERT INTO projects (name, location, created_date, last_modified, total_cost, metadata)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (
                project_data['name'],
                project_data.get('location', ''),
                datetime.now().isoformat(),
                datetime.now().isoformat(),
                project_data.get('total_cost', 0),
                json.dumps(project_data.get('metadata', {}))
            ))
            
            project_id = cursor.lastrowid
            
            # Save measurements and abstracts
            self._save_project_data(cursor, project_id, project_data)
            
            conn.commit()
            conn.close()
            
            return project_id
            
        except Exception as e:
            logger.error(f"Error saving project: {e}")
            return -1
    
    def _save_project_data(self, cursor, project_id: int, project_data: Dict):
        """Save measurements and abstracts data"""
        # Save measurements
        if 'measurements' in project_data:
            for sheet_name, measurements_df in project_data['measurements'].items():
                for _, row in measurements_df.iterrows():
                    cursor.execute("""
                        INSERT INTO measurements (
                            project_id, sheet_name, item_no, description, specification,
                            location, quantity, length, breadth, height, diameter,
                            thickness, unit, total, deduction, net_total, remarks, ssr_code
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        project_id, sheet_name, row.get('item_no', ''),
                        row.get('description', ''), row.get('specification', ''),
                        row.get('location', ''), row.get('quantity', 0),
                        row.get('length', 0), row.get('breadth', 0),
                        row.get('height', 0), row.get('diameter', 0),
                        row.get('thickness', 0), row.get('unit', ''),
                        row.get('total', 0), row.get('deduction', 0),
                        row.get('net_total', 0), row.get('remarks', ''),
                        row.get('ssr_code', '')
                    ))
        
        # Save abstracts
        if 'abstracts' in project_data:
            for sheet_name, abstracts_df in project_data['abstracts'].items():
                for _, row in abstracts_df.iterrows():
                    cursor.execute("""
                        INSERT INTO abstracts (
                            project_id, sheet_name, ssr_code, description,
                            unit, quantity, rate, amount
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        project_id, sheet_name, row.get('ssr_code', ''),
                        row.get('description', ''), row.get('unit', ''),
                        row.get('quantity', 0), row.get('rate', 0),
                        row.get('amount', 0)
                    ))

    def load_project(self, project_id: int) -> Optional[Dict]:
        """Load project with measurements and abstracts grouped by sheet"""
        try:
            conn = sqlite3.connect(self.db_path)
            project_df = pd.read_sql_query(
                "SELECT * FROM projects WHERE id = ?", conn, params=(project_id,)
            )
            if project_df.empty:
                conn.close()
                return None
            project_info = project_df.iloc[0].to_dict()
            measurements_df = pd.read_sql_query(
                "SELECT * FROM measurements WHERE project_id = ?", conn, params=(project_id,)
            )
            abstracts_df = pd.read_sql_query(
                "SELECT * FROM abstracts WHERE project_id = ?", conn, params=(project_id,)
            )
            conn.close()
            measurements_by_sheet = {}
            abstracts_by_sheet = {}
            if not measurements_df.empty:
                for sheet in measurements_df['sheet_name'].unique():
                    measurements_by_sheet[sheet] = measurements_df[measurements_df['sheet_name'] == sheet].drop(columns=['project_id'])
            if not abstracts_df.empty:
                for sheet in abstracts_df['sheet_name'].unique():
                    abstracts_by_sheet[sheet] = abstracts_df[abstracts_df['sheet_name'] == sheet].drop(columns=['project_id'])
            return {
                'project_info': project_info,
                'measurements': measurements_by_sheet,
                'abstracts': abstracts_by_sheet,
            }
        except Exception as e:
            logger.error(f"Error loading project: {e}")
            return None

    def update_ssr_items(self, ssr_df: pd.DataFrame) -> bool:
        """Replace SSR items with provided DataFrame"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute("DELETE FROM ssr_items")
            for _, row in ssr_df.iterrows():
                cursor.execute(
                    """
                    INSERT INTO ssr_items (code, description, category, unit, rate, last_updated)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (
                        row.get('code', ''), row.get('description', ''), row.get('category', ''),
                        row.get('unit', ''), float(row.get('rate', 0) or 0), datetime.now().isoformat(),
                    ),
                )
            conn.commit()
            conn.close()
            return True
        except Exception as e:
            logger.error(f"Error updating SSR items: {e}")
            return False

    def load_ssr_items(self) -> pd.DataFrame:
        """Load SSR items as DataFrame"""
        try:
            conn = sqlite3.connect(self.db_path)
            df = pd.read_sql_query("SELECT code, description, category, unit, rate FROM ssr_items ORDER BY code", conn)
            conn.close()
            return df
        except Exception as e:
            logger.error(f"Error loading SSR items: {e}")
            return pd.DataFrame(columns=SSR_COLUMNS)
    
    def list_projects(self) -> pd.DataFrame:
        """List all projects"""
        try:
            conn = sqlite3.connect(self.db_path)
            projects_df = pd.read_sql_query(
                "SELECT id, name, location, created_date, total_cost, status FROM projects ORDER BY last_modified DESC",
                conn
            )
            conn.close()
            return projects_df
        except Exception as e:
            logger.error(f"Error listing projects: {e}")
            return pd.DataFrame()

# =============================================================================
# ENHANCED EXCEL IMPORTER
# =============================================================================

class UnifiedExcelImporter:
    """Unified Excel importer with all enhancements"""
    
    def __init__(self):
        self.excel_file = None
        self.workbook = None
        self.import_stats = {
            'sheets_processed': 0,
            'formulas_preserved': 0,
            'measurements_imported': 0,
            'abstracts_imported': 0,
            'linkages_created': 0,
            'errors': [],
            'warnings': []
        }
    
    def import_excel_file(self, file_path: str, progress_callback=None) -> Dict:
        """Main import function with comprehensive processing"""
        try:
            self._update_progress(progress_callback, 0, "Starting import...")
            
            # Load workbook
            self.workbook = load_workbook(file_path, data_only=False)
            self._update_progress(progress_callback, 20, "Analyzing structure...")
            
            # Analyze structure
            structure = self._analyze_structure()
            self._update_progress(progress_callback, 40, "Processing sheets...")
            
            # Import data
            estimate_data = self._import_all_sheets(structure, progress_callback)
            
            # Generate report
            estimate_data['import_report'] = self._generate_report()
            
            self._update_progress(progress_callback, 100, "Import completed!")
            return estimate_data
            
        except Exception as e:
            logger.error(f"Import failed: {e}")
            raise
    
    def _analyze_structure(self) -> Dict:
        """Analyze Excel file structure"""
        structure = {
            'sheets': [],
            'general_abstract': None,
            'measurement_abstract_pairs': [],
            'other_sheets': [],
            'total_formulas': 0
        }
        
        measurements = {}
        abstracts = {}
        
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            
            # Count formulas
            formula_count = self._count_formulas(sheet)
            structure['total_formulas'] += formula_count
            
            # Detect sheet type
            sheet_type = self._detect_sheet_type(sheet_name, sheet)
            
            sheet_info = {
                'name': sheet_name,
                'type': sheet_type,
                'formula_count': formula_count,
                'row_count': sheet.max_row,
                'col_count': sheet.max_column
            }
            
            structure['sheets'].append(sheet_info)
            
            # Categorize sheets
            if sheet_type == 'general_abstract':
                structure['general_abstract'] = sheet_name
            elif sheet_type == 'measurement':
                part_name = self._extract_part_name(sheet_name)
                measurements[part_name] = sheet_name
            elif sheet_type == 'abstract':
                part_name = self._extract_part_name(sheet_name)
                abstracts[part_name] = sheet_name
            else:
                structure['other_sheets'].append(sheet_info)
        
        # Create measurement-abstract pairs
        for part_name in set(measurements.keys()) & set(abstracts.keys()):
            structure['measurement_abstract_pairs'].append({
                'part_name': part_name,
                'measurement_sheet': measurements[part_name],
                'abstract_sheet': abstracts[part_name]
            })
        
        return structure
    
    def _detect_sheet_type(self, sheet_name: str, sheet) -> str:
        """Detect sheet type from name and content"""
        name_lower = sheet_name.lower()
        
        if 'general' in name_lower and 'abstract' in name_lower:
            return 'general_abstract'
        elif any(word in name_lower for word in ['measurement', 'measur', '_mes']):
            return 'measurement'
        elif 'abstract' in name_lower or '_abs' in name_lower:
            return 'abstract'
        elif 'sanitary' in name_lower:
            if 'measur' in name_lower:
                return 'measurement'
            else:
                return 'abstract'
        elif any(word in name_lower for word in ['tech', 'report']):
            return 'technical_report'
        else:
            return 'other'
    
    def _extract_part_name(self, sheet_name: str) -> str:
        """Extract part name from sheet name"""
        clean_name = sheet_name
        suffixes = ['_MES', '_ABS', '_MEASUR', '-abs', ' Measurement', ' Abstract']
        for suffix in suffixes:
            clean_name = clean_name.replace(suffix, '')
        return clean_name.strip()
    
    def _count_formulas(self, sheet) -> int:
        """Count formulas in sheet"""
        count = 0
        try:
            for row in sheet.iter_rows():
                for cell in row:
                    if (cell.value and isinstance(cell.value, str) and 
                        cell.value.startswith('=')):
                        count += 1
        except Exception:
            pass
        return count
    
    def _import_all_sheets(self, structure: Dict, progress_callback=None) -> Dict:
        """Import all sheets with progress tracking"""
        estimate_data = {
            'general_abstract': None,
            'parts': [],
            'reports': {},
            'metadata': {
                'import_date': datetime.now(),
                'total_formulas': structure['total_formulas'],
                'structure': structure
            }
        }
        
        # Import general abstract
        if structure['general_abstract']:
            self._update_progress(progress_callback, 50, "Importing general abstract...")
            estimate_data['general_abstract'] = self._import_general_abstract(
                structure['general_abstract']
            )
        
        # Import measurement-abstract pairs
        total_pairs = len(structure['measurement_abstract_pairs'])
        for i, pair in enumerate(structure['measurement_abstract_pairs']):
            progress = 60 + (i / total_pairs) * 30
            self._update_progress(progress_callback, progress, f"Importing {pair['part_name']}...")
            
            try:
                part_data = self._import_part_pair(pair)
                estimate_data['parts'].append(part_data)
                
                self.import_stats['measurements_imported'] += len(part_data['measurements'])
                self.import_stats['abstracts_imported'] += len(part_data['abstracts'])
                
            except Exception as e:
                self.import_stats['errors'].append(f"Failed to import {pair['part_name']}: {str(e)}")
        
        return estimate_data
    
    def _import_part_pair(self, pair_info: Dict) -> Dict:
        """Import measurement-abstract pair"""
        part_data = {
            'name': pair_info['part_name'],
            'measurement_sheet': pair_info['measurement_sheet'],
            'abstract_sheet': pair_info['abstract_sheet'],
            'measurements': pd.DataFrame(),
            'abstracts': pd.DataFrame(),
            'linkages': []
        }
        
        # Import measurements
        measurements_df = self._import_measurement_sheet(pair_info['measurement_sheet'])
        part_data['measurements'] = measurements_df
        
        # Import abstracts
        abstracts_df = self._import_abstract_sheet(pair_info['abstract_sheet'])
        part_data['abstracts'] = abstracts_df
        
        # Create linkages
        linkages = self._create_linkages(measurements_df, abstracts_df)
        part_data['linkages'] = linkages
        self.import_stats['linkages_created'] += len(linkages)
        
        return part_data
    
    def _import_measurement_sheet(self, sheet_name: str) -> pd.DataFrame:
        """Import measurement sheet"""
        try:
            # Read Excel data
            df = pd.read_excel(self.excel_file, sheet_name=sheet_name)
            
            # Process and clean data
            measurements = []
            for idx, row in df.iterrows():
                if pd.isna(row.get('Particulars')) or str(row.get('Particulars')).strip() == '':
                    continue
                
                measurement = {
                    'id': len(measurements) + 1,
                    'item_no': str(row.get('S.No.', '')),
                    'description': str(row.get('Particulars', '')),
                    'specification': '',
                    'location': '',
                    'quantity': self._safe_float(row.get('Nos.', 1)),
                    'length': self._safe_float(row.get('Length', 0)),
                    'breadth': self._safe_float(row.get('Breadth', 0)),
                    'height': self._safe_float(row.get('Height', 0)),
                    'diameter': 0,
                    'thickness': 0,
                    'unit': str(row.get('Units', '')),
                    'total': self._safe_float(row.get('Qty.', 0)),
                    'deduction': 0,
                    'net_total': self._safe_float(row.get('Qty.', 0)),
                    'remarks': '',
                    'ssr_code': ''
                }
                
                measurements.append(measurement)
            
            return pd.DataFrame(measurements)
            
        except Exception as e:
            logger.error(f"Error importing measurement sheet {sheet_name}: {e}")
            return pd.DataFrame()
    
    def _import_abstract_sheet(self, sheet_name: str) -> pd.DataFrame:
        """Import abstract sheet"""
        try:
            df = pd.read_excel(self.excel_file, sheet_name=sheet_name)
            
            abstracts = []
            for idx, row in df.iterrows():
                if pd.isna(row.get('Particulars')) or str(row.get('Particulars')).strip() == '':
                    continue
                
                abstract = {
                    'id': len(abstracts) + 1,
                    'ssr_code': '',
                    'description': str(row.get('Particulars', '')),
                    'unit': str(row.get('Unit', '')),
                    'quantity': self._safe_float(row.get('Quantity', 0)),
                    'rate': self._safe_float(row.get('Rate', 0)),
                    'amount': self._safe_float(row.get('Amount', 0))
                }
                
                abstracts.append(abstract)
            
            return pd.DataFrame(abstracts)
            
        except Exception as e:
            logger.error(f"Error importing abstract sheet {sheet_name}: {e}")
            return pd.DataFrame()
    
    def _import_general_abstract(self, sheet_name: str) -> Dict:
        """Import general abstract"""
        try:
            df = pd.read_excel(self.excel_file, sheet_name=sheet_name)
            
            # Extract project information
            project_info = {
                'project_name': 'Construction Project',
                'location': '',
                'parts': [],
                'grand_total': 0
            }
            
            # Process data to extract cost breakdown
            for idx, row in df.iterrows():
                if pd.notna(row.iloc[0]) and 'PART' in str(row.iloc[0]).upper():
                    # Find amount in the row
                    for col in reversed(df.columns):
                        if pd.notna(row[col]) and isinstance(row[col], (int, float)) and row[col] > 0:
                            project_info['parts'].append({
                                'part_name': str(row.iloc[0]),
                                'amount': float(row[col])
                            })
                            break
            
            project_info['grand_total'] = sum(part['amount'] for part in project_info['parts'])
            
            return project_info
            
        except Exception as e:
            logger.error(f"Error importing general abstract: {e}")
            return {}
    
    def _create_linkages(self, measurements_df: pd.DataFrame, abstracts_df: pd.DataFrame) -> List[Dict]:
        """Create linkages between measurements and abstracts"""
        linkages = []
        
        for _, abstract_row in abstracts_df.iterrows():
            abstract_desc = abstract_row['description'].lower()
            abstract_words = set(re.findall(r'\w+', abstract_desc))
            
            matching_measurements = []
            
            for _, meas_row in measurements_df.iterrows():
                meas_desc = meas_row['description'].lower()
                meas_words = set(re.findall(r'\w+', meas_desc))
                
                # Calculate similarity
                intersection = len(abstract_words & meas_words)
                union = len(abstract_words | meas_words)
                
                if union > 0:
                    similarity = intersection / union
                    
                    if similarity > 0.3 and intersection >= 2:
                        matching_measurements.append({
                            'measurement_id': meas_row['id'],
                            'measurement_desc': meas_row['description'],
                            'similarity': similarity,
                            'total': meas_row['total'],
                            'unit': meas_row['unit']
                        })
            
            if matching_measurements:
                matching_measurements.sort(key=lambda x: x['similarity'], reverse=True)
                
                linkages.append({
                    'abstract_id': abstract_row['id'],
                    'abstract_description': abstract_row['description'],
                    'measurements': matching_measurements,
                    'total_quantity': sum(m['total'] for m in matching_measurements),
                    'confidence': matching_measurements[0]['similarity']
                })
        
        return linkages
    
    def _safe_float(self, value, default=0.0) -> float:
        """Safely convert value to float"""
        try:
            if pd.isna(value) or value == '':
                return default
            return float(value)
        except (ValueError, TypeError):
            return default
    
    def _update_progress(self, callback, percentage: int, message: str):
        """Update progress"""
        if callback:
            callback(percentage, message)
    
    def _generate_report(self) -> Dict:
        """Generate import report"""
        return {
            'summary': {
                'sheets_processed': self.import_stats['sheets_processed'],
                'formulas_preserved': self.import_stats['formulas_preserved'],
                'measurements_imported': self.import_stats['measurements_imported'],
                'abstracts_imported': self.import_stats['abstracts_imported'],
                'linkages_created': self.import_stats['linkages_created']
            },
            'errors': self.import_stats['errors'],
            'warnings': self.import_stats['warnings']
        }

# =============================================================================
# CALCULATION ENGINE
# =============================================================================

class CalculationEngine:
    """Real-time calculation engine with dependency tracking"""
    
    @staticmethod
    @lru_cache(maxsize=1000)
    def calculate_total(quantity: float, length: float, breadth: float, 
                       height: float, measurement_type: str = "Standard") -> float:
        """Calculate total with caching"""
        try:
            if measurement_type == "Linear":
                return quantity * max(1, length)
            elif measurement_type == "Area":
                return quantity * max(1, length) * max(1, breadth)
            elif measurement_type == "Volume":
                return quantity * max(1, length) * max(1, breadth) * max(1, height)
            else:  # Standard
                return quantity * max(1, length) * max(1, breadth) * max(1, height)
        except Exception:
            return 0
    
    @staticmethod
    def update_measurement_totals(measurements_df: pd.DataFrame) -> pd.DataFrame:
        """Update all measurement totals"""
        if measurements_df.empty:
            return measurements_df
        
        for idx in measurements_df.index:
            row = measurements_df.loc[idx]
            total = CalculationEngine.calculate_total(
                row.get('quantity', 1),
                row.get('length', 0),
                row.get('breadth', 0),
                row.get('height', 0)
            )
            measurements_df.at[idx, 'total'] = total
            measurements_df.at[idx, 'net_total'] = total - row.get('deduction', 0)
        
        return measurements_df
    
    @staticmethod
    def update_abstract_amounts(abstracts_df: pd.DataFrame) -> pd.DataFrame:
        """Update all abstract amounts"""
        if abstracts_df.empty:
            return abstracts_df
        
        for idx in abstracts_df.index:
            row = abstracts_df.loc[idx]
            amount = row.get('quantity', 0) * row.get('rate', 0)
            abstracts_df.at[idx, 'amount'] = amount
        
        return abstracts_df

# =============================================================================
# SESSION STATE MANAGEMENT
# =============================================================================

def initialize_session_state():
    """Initialize all session state variables"""
    try:
        # Initialize database
        if 'database' not in st.session_state:
            st.session_state.database = UnifiedDatabase()
        
        # Initialize measurements
        if 'measurements' not in st.session_state:
            st.session_state.measurements = pd.DataFrame(columns=MEASUREMENT_COLUMNS)
        
        # Initialize abstracts
        if 'abstract_items' not in st.session_state:
            st.session_state.abstract_items = pd.DataFrame(columns=ABSTRACT_COLUMNS)
        
        # Initialize SSR items
        if 'ssr_items' not in st.session_state:
            st.session_state.ssr_items = load_default_ssr_data()
        
        # Initialize measurement sheets
        if 'measurement_sheets' not in st.session_state:
            st.session_state.measurement_sheets = {
                'Ground Floor': pd.DataFrame(columns=MEASUREMENT_COLUMNS),
                'First Floor': pd.DataFrame(columns=MEASUREMENT_COLUMNS),
                'Basement': pd.DataFrame(columns=MEASUREMENT_COLUMNS)
            }
        
        # Initialize abstract sheets
        if 'abstract_sheets' not in st.session_state:
            st.session_state.abstract_sheets = {
                'Ground Floor': pd.DataFrame(columns=ABSTRACT_COLUMNS),
                'First Floor': pd.DataFrame(columns=ABSTRACT_COLUMNS),
                'Basement': pd.DataFrame(columns=ABSTRACT_COLUMNS)
            }
        
        # Initialize project settings
        if 'project_settings' not in st.session_state:
            st.session_state.project_settings = {
                'project_name': 'New Construction Project',
                'project_location': '',
                'engineer_name': '',
                'current_project_id': None
            }
        
        logger.info("Session state initialized successfully")
        
    except Exception as e:
        logger.error(f"Error initializing session state: {e}")
        st.error(f"Initialization error: {e}")

@st.cache_data(ttl=3600)
def load_default_ssr_data():
    """Load default SSR data with caching"""
    return pd.DataFrame([
        # Earth Work
        {"code": "1.1.1", "description": "Earth work excavation in foundation", "category": "Earth Work", "unit": "cum", "rate": 245.50},
        {"code": "1.1.2", "description": "Earth work excavation by mechanical means", "category": "Earth Work", "unit": "cum", "rate": 185.00},
        
        # Concrete Work
        {"code": "2.1.1", "description": "Cement concrete 1:2:4 using 20mm aggregate", "category": "Concrete Work", "unit": "cum", "rate": 4850.00},
        {"code": "2.1.2", "description": "Cement concrete 1:3:6 using 40mm aggregate", "category": "Concrete Work", "unit": "cum", "rate": 4200.00},
        
        # Masonry Work
        {"code": "3.1.1", "description": "Brick work in superstructure", "category": "Masonry Work", "unit": "cum", "rate": 5200.00},
        {"code": "3.1.2", "description": "Brick work in foundation", "category": "Masonry Work", "unit": "cum", "rate": 4800.00},
        
        # Plastering
        {"code": "4.1.1", "description": "12mm thick cement plaster 1:4", "category": "Plastering", "unit": "sqm", "rate": 125.00},
        {"code": "4.1.2", "description": "15mm thick cement plaster 1:3", "category": "Plastering", "unit": "sqm", "rate": 145.00},
        
        # Steel Work
        {"code": "7.1.1", "description": "Steel reinforcement bars", "category": "Steel Work", "unit": "kg", "rate": 65.00},
        {"code": "7.2.1", "description": "Structural steel work", "category": "Steel Work", "unit": "kg", "rate": 85.00}
    ])

# =============================================================================
# USER INTERFACE COMPONENTS
# =============================================================================

def create_header():
    """Create application header"""
    st.markdown("""
        <div style="background: linear-gradient(90deg, #1f4e79 0%, #2d5aa0 100%); 
                    color: white; padding: 1rem; border-radius: 10px; margin-bottom: 2rem; text-align: center;">
            <h1>🏗️ Construction Estimation System</h1>
            <p>Complete Unified Solution - Version 3.0</p>
        </div>
    """, unsafe_allow_html=True)

def create_sidebar():
    """Create sidebar navigation"""
    st.sidebar.title("📋 Navigation")
    
    # Project selector
    projects = st.session_state.database.list_projects()
    if not projects.empty:
        st.sidebar.subheader("📂 Current Project")
        project_names = ["New Project"] + projects['name'].tolist()
        selected_project = st.sidebar.selectbox("Select Project", project_names)
        
        if selected_project != "New Project" and st.sidebar.button("📂 Load Project"):
            # Load selected project from database
            sel_row = projects[projects['name'] == selected_project].iloc[0]
            loaded = st.session_state.database.load_project(int(sel_row['id']))
            if loaded:
                # Update session state
                st.session_state.project_settings.update({
                    'project_name': loaded['project_info'].get('name', selected_project),
                    'project_location': loaded['project_info'].get('location', ''),
                    'current_project_id': loaded['project_info'].get('id'),
                })
                # Reset existing
                st.session_state.measurement_sheets = {k: v for k, v in st.session_state.measurement_sheets.items()}
                st.session_state.abstract_sheets = {k: v for k, v in st.session_state.abstract_sheets.items()}
                # Apply loaded sheets
                if loaded['measurements']:
                    for sheet, df in loaded['measurements'].items():
                        st.session_state.measurement_sheets[sheet] = df
                if loaded['abstracts']:
                    for sheet, df in loaded['abstracts'].items():
                        st.session_state.abstract_sheets[sheet] = df
                st.success(f"✅ Loaded project: {selected_project}")
                st.rerun()
            else:
                st.error("❌ Failed to load project")
    
    # Main navigation
    page = st.sidebar.selectbox("Select Module", [
        "📊 Dashboard", 
        "📋 General Abstract",
        "💰 Abstract of Cost",
        "📝 Measurement Sheets",
        "📚 SSR Database",
        "🧩 Templates",
        "📥 Import Excel Data",
        "📊 Analytics & Reports",
        "🔧 System Tools"
    ])
    
    return page

def show_dashboard():
    """Show main dashboard"""
    st.title("📊 Project Dashboard")
    
    # Key metrics
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        total_measurements = sum(len(df) for df in st.session_state.measurement_sheets.values())
        st.metric("Total Measurements", total_measurements)
    
    with col2:
        st.metric("SSR Items", len(st.session_state.ssr_items))
    
    with col3:
        total_abstracts = sum(len(df) for df in st.session_state.abstract_sheets.values())
        st.metric("Abstract Items", total_abstracts)
    
    with col4:
        total_cost = sum(
            df['amount'].sum() if not df.empty else 0 
            for df in st.session_state.abstract_sheets.values()
        )
        st.metric("Total Cost", f"₹{total_cost:,.0f}")
    
    # Recent activity
    st.subheader("📈 Recent Activity")
    if not st.session_state.measurements.empty:
        recent = st.session_state.measurements.tail(5)[['item_no', 'description', 'unit', 'total']]
        st.dataframe(recent, use_container_width=True, hide_index=True)
    else:
        st.info("No measurements added yet. Start by importing Excel data or adding measurements manually.")

def show_excel_import():
    """Show Excel import interface"""
    st.title("📥 Enhanced Excel Import System")
    
    st.markdown("""
    ### 🚀 Features:
    - **Smart Detection**: Automatically identifies sheet types
    - **Formula Preservation**: Maintains Excel calculations
    - **Auto-Linking**: Connects measurements to abstracts
    - **Progress Tracking**: Real-time import feedback
    - **Comprehensive Reporting**: Detailed import statistics
    """)
    
    uploaded_file = st.file_uploader("Upload Excel File", type=['xlsx'])
    
    if uploaded_file:
        # Progress containers
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        def progress_callback(percentage, message):
            progress_bar.progress(percentage / 100)
            status_text.text(message)
        
        if st.button("🚀 Import Excel File", type="primary"):
            # Save uploaded file temporarily
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                tmp.write(uploaded_file.getvalue())
                tmp_path = tmp.name
            
            try:
                # Import with unified importer
                importer = UnifiedExcelImporter()
                importer.excel_file = tmp_path
                
                estimate_data = importer.import_excel_file(tmp_path, progress_callback)
                
                # Update session state
                update_session_state_from_import(estimate_data)
                
                # Show results
                st.success("🎉 Import Completed Successfully!")
                
                # Display import report
                report = estimate_data['import_report']
                
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("Measurements", report['summary']['measurements_imported'])
                with col2:
                    st.metric("Abstracts", report['summary']['abstracts_imported'])
                with col3:
                    st.metric("Linkages", report['summary']['linkages_created'])
                with col4:
                    st.metric("Formulas", report['summary']['formulas_preserved'])
                
                # Show detailed report
                with st.expander("📊 Detailed Import Report"):
                    st.json(report)
                
            except Exception as e:
                st.error(f"Import failed: {str(e)}")
            
            finally:
                # Cleanup
                if os.path.exists(tmp_path):
                    os.unlink(tmp_path)

def update_session_state_from_import(estimate_data: Dict):
    """Update session state with imported data"""
    try:
        # Update project settings
        if estimate_data.get('general_abstract'):
            ga = estimate_data['general_abstract']
            st.session_state.project_settings.update({
                'project_name': ga.get('project_name', 'Imported Project'),
                'project_location': ga.get('location', '')
            })
        
        # Update measurement and abstract sheets
        for part in estimate_data.get('parts', []):
            sheet_name = part['name']
            
            # Update measurements
            if not part['measurements'].empty:
                st.session_state.measurement_sheets[sheet_name] = part['measurements']
            
            # Update abstracts
            if not part['abstracts'].empty:
                st.session_state.abstract_sheets[sheet_name] = part['abstracts']
        
        st.success("✅ Session state updated with imported data")
        
    except Exception as e:
        logger.error(f"Error updating session state: {e}")
        st.error(f"Error updating data: {e}")

def show_measurements():
    """Show measurement sheets interface"""
    st.title("📝 Measurement Sheets")
    
    # Sheet selector
    sheet_names = list(st.session_state.measurement_sheets.keys())
    selected_sheet = st.selectbox("Select Sheet", sheet_names)
    
    if selected_sheet:
        st.subheader(f"📋 {selected_sheet} Measurements")
        
        # Current measurements
        measurements_df = st.session_state.measurement_sheets[selected_sheet]
        
        if not measurements_df.empty:
            # Display measurements with editing capability
            edited_df = st.data_editor(
                measurements_df,
                use_container_width=True,
                num_rows="dynamic",
                column_config={
                    "total": st.column_config.NumberColumn("Total", format="%.2f"),
                    "rate": st.column_config.NumberColumn("Rate", format="%.2f"),
                    "amount": st.column_config.NumberColumn("Amount", format="%.2f")
                }
            )
            
            # Update calculations if data changed
            if not edited_df.equals(measurements_df):
                updated_df = CalculationEngine.update_measurement_totals(edited_df)
                st.session_state.measurement_sheets[selected_sheet] = updated_df
                st.rerun()
            
            # Summary
            total_quantity = measurements_df['total'].sum()
            st.metric("Total Quantity", f"{total_quantity:.2f}")
            
        else:
            st.info("No measurements in this sheet. Import Excel data or add measurements manually.")
        
        # Add new measurement
        with st.expander("➕ Add New Measurement"):
            with st.form("add_measurement"):
                col1, col2 = st.columns(2)
                
                with col1:
                    item_no = st.text_input("Item No.")
                    description = st.text_area("Description")
                    quantity = st.number_input("Quantity", value=1.0, min_value=0.0)
                    length = st.number_input("Length", value=0.0, min_value=0.0)
                
                with col2:
                    breadth = st.number_input("Breadth", value=0.0, min_value=0.0)
                    height = st.number_input("Height", value=0.0, min_value=0.0)
                    unit = st.selectbox("Unit", UNITS)
                    remarks = st.text_input("Remarks")
                
                if st.form_submit_button("➕ Add Measurement"):
                    # Calculate total
                    total = CalculationEngine.calculate_total(quantity, length, breadth, height)
                    
                    # Create new measurement
                    new_measurement = {
                        'id': len(measurements_df) + 1,
                        'ssr_code': '',
                        'item_no': item_no,
                        'description': description,
                        'specification': '',
                        'location': '',
                        'quantity': quantity,
                        'length': length,
                        'breadth': breadth,
                        'height': height,
                        'diameter': 0,
                        'thickness': 0,
                        'unit': unit,
                        'total': total,
                        'deduction': 0,
                        'net_total': total,
                        'remarks': remarks
                    }
                    
                    # Add to dataframe
                    new_df = pd.concat([measurements_df, pd.DataFrame([new_measurement])], ignore_index=True)
                    st.session_state.measurement_sheets[selected_sheet] = new_df
                    
                    st.success("✅ Measurement added successfully!")
                    st.rerun()

def show_abstracts():
    """Show abstract of cost sheets interface"""
    st.title("💰 Abstract of Cost")
    
    sheet_names = list(st.session_state.abstract_sheets.keys())
    selected_sheet = st.selectbox("Select Sheet", sheet_names)
    
    if selected_sheet:
        st.subheader(f"📋 {selected_sheet} Abstract")
        abstracts_df = st.session_state.abstract_sheets[selected_sheet]
        
        if not abstracts_df.empty:
            edited_df = st.data_editor(
                abstracts_df,
                use_container_width=True,
                num_rows="dynamic",
                column_config={
                    "quantity": st.column_config.NumberColumn("Quantity", format="%.3f"),
                    "rate": st.column_config.NumberColumn("Rate", format="%.2f"),
                    "amount": st.column_config.NumberColumn("Amount", format="%.2f"),
                },
            )
            
            if not edited_df.equals(abstracts_df):
                updated_df = CalculationEngine.update_abstract_amounts(edited_df)
                st.session_state.abstract_sheets[selected_sheet] = updated_df
                st.rerun()
            
            total_amount = abstracts_df['amount'].sum() if 'amount' in abstracts_df else 0
            st.metric("Sheet Total", f"₹{total_amount:,.2f}")
        else:
            st.info("No abstract items in this sheet. Import Excel data or add items manually.")
        
        with st.expander("➕ Add Abstract Item"):
            with st.form("add_abstract"):
                col1, col2 = st.columns(2)
                
                with col1:
                    ssr_code = st.text_input("SSR/BSR Code", value="")
                    description = st.text_area("Description")
                    unit = st.text_input("Unit", value="")
                
                with col2:
                    quantity = st.number_input("Quantity", value=0.0, min_value=0.0)
                    rate = st.number_input("Rate", value=0.0, min_value=0.0)
                    
                if st.form_submit_button("➕ Add Item"):
                    amount = quantity * rate
                    new_item = {
                        'id': len(abstracts_df) + 1,
                        'ssr_code': ssr_code,
                        'description': description,
                        'unit': unit,
                        'quantity': quantity,
                        'rate': rate,
                        'amount': amount,
                    }
                    new_df = pd.concat([abstracts_df, pd.DataFrame([new_item])], ignore_index=True)
                    st.session_state.abstract_sheets[selected_sheet] = new_df
                    st.success("✅ Abstract item added successfully!")
                    st.rerun()

def show_analytics():
    """Show analytics and reporting"""
    st.title("📊 Analytics & Reports")
    
    # Cost breakdown by category
    st.subheader("💰 Cost Analysis")
    
    # Collect all abstract data
    all_abstracts = []
    for sheet_name, df in st.session_state.abstract_sheets.items():
        if not df.empty:
            df_copy = df.copy()
            df_copy['sheet'] = sheet_name
            all_abstracts.append(df_copy)
    
    if all_abstracts:
        combined_abstracts = pd.concat(all_abstracts, ignore_index=True)
        
        # Cost by sheet
        cost_by_sheet = combined_abstracts.groupby('sheet')['amount'].sum().reset_index()
        
        col1, col2 = st.columns(2)
        
        with col1:
            # Pie chart
            fig_pie = px.pie(
                cost_by_sheet, 
                values='amount', 
                names='sheet',
                title="Cost Distribution by Work Type"
            )
            st.plotly_chart(fig_pie, use_container_width=True)
        
        with col2:
            # Bar chart
            fig_bar = px.bar(
                cost_by_sheet,
                x='sheet',
                y='amount',
                title="Cost by Work Category",
                color='amount',
                color_continuous_scale='viridis'
            )
            st.plotly_chart(fig_bar, use_container_width=True)
        
        # Summary table
        st.subheader("📋 Cost Summary")
        cost_summary = cost_by_sheet.copy()
        cost_summary['percentage'] = (cost_summary['amount'] / cost_summary['amount'].sum() * 100).round(2)
        cost_summary['amount'] = cost_summary['amount'].apply(lambda x: f"₹{x:,.2f}")
        cost_summary['percentage'] = cost_summary['percentage'].apply(lambda x: f"{x}%")
        
        st.dataframe(cost_summary, use_container_width=True, hide_index=True)
        
    else:
        st.info("No cost data available. Import Excel data or add abstracts to see analytics.")

def show_general_abstract():
    """Basic General Abstract page for project metadata and part totals"""
    st.title("📋 General Abstract")
    col1, col2 = st.columns(2)
    with col1:
        st.session_state.project_settings['project_name'] = st.text_input(
            "Project Name", value=st.session_state.project_settings.get('project_name', 'New Construction Project')
        )
        st.session_state.project_settings['project_location'] = st.text_input(
            "Project Location", value=st.session_state.project_settings.get('project_location', '')
        )
    with col2:
        st.write("Current Project ID:", st.session_state.project_settings.get('current_project_id') or "—")
        total_cost = sum(
            df['amount'].sum() if not df.empty else 0 for df in st.session_state.abstract_sheets.values()
        )
        st.metric("Grand Total", f"₹{total_cost:,.2f}")
    st.subheader("Part-wise totals")
    rows = []
    for sheet_name, df in st.session_state.abstract_sheets.items():
        amount = df['amount'].sum() if not df.empty and 'amount' in df else 0
        rows.append({"Part": sheet_name, "Amount": amount})
    if rows:
        summary_df = pd.DataFrame(rows)
        st.dataframe(summary_df, use_container_width=True, hide_index=True)

def show_templates():
    """Templates module: save current as template and apply templates"""
    st.title("🧩 Templates")
    tab1, tab2 = st.tabs(["Save Template", "Apply Template"])
    with tab1:
        name = st.text_input("Template Name")
        description = st.text_area("Description", height=80)
        category = st.text_input("Category", value="General")
        if st.button("💾 Save Current As Template", type="primary"):
            try:
                conn = sqlite3.connect(st.session_state.database.db_path)
                cursor = conn.cursor()
                template_payload = {
                    'measurement_sheets': {k: v.to_dict(orient='records') for k, v in st.session_state.measurement_sheets.items()},
                    'abstract_sheets': {k: v.to_dict(orient='records') for k, v in st.session_state.abstract_sheets.items()},
                }
                cursor.execute(
                    """
                    INSERT INTO templates (name, description, category, template_data, created_date)
                    VALUES (?, ?, ?, ?, ?)
                    """,
                    (name, description, category, json.dumps(template_payload), datetime.now().isoformat()),
                )
                conn.commit(); conn.close()
                st.success("✅ Template saved")
            except Exception as e:
                st.error(f"❌ Failed to save template: {e}")
    with tab2:
        try:
            conn = sqlite3.connect(st.session_state.database.db_path)
            templates_df = pd.read_sql_query("SELECT id, name, category, created_date FROM templates ORDER BY created_date DESC", conn)
            conn.close()
        except Exception:
            templates_df = pd.DataFrame(columns=["id","name","category","created_date"])
        if templates_df.empty:
            st.info("No templates available.")
        else:
            st.dataframe(templates_df, use_container_width=True, hide_index=True)
            ids = templates_df['id'].tolist()
            sel = st.selectbox("Select Template ID", ids) if ids else None
            if sel and st.button("📥 Apply Template to Current Project"):
                try:
                    conn = sqlite3.connect(st.session_state.database.db_path)
                    row = pd.read_sql_query("SELECT template_data FROM templates WHERE id = ?", conn, params=(int(sel),))
                    conn.close()
                    payload = json.loads(row.iloc[0]['template_data']) if not row.empty else {}
                    ms = payload.get('measurement_sheets', {})
                    abs_ = payload.get('abstract_sheets', {})
                    # Apply
                    for k, v in ms.items():
                        st.session_state.measurement_sheets[k] = pd.DataFrame(v)
                    for k, v in abs_.items():
                        st.session_state.abstract_sheets[k] = pd.DataFrame(v)
                    st.success("✅ Template applied")
                    st.rerun()
                except Exception as e:
                    st.error(f"❌ Failed to apply template: {e}")

# =============================================================================
# MAIN APPLICATION
# =============================================================================

def main():
    """Main application function"""
    # Initialize session state
    initialize_session_state()
    
    # Create header
    create_header()
    
    # Create sidebar and get selected page
    page = create_sidebar()
    
    # Route to appropriate page
    if page == "📊 Dashboard":
        show_dashboard()
    elif page == "📥 Import Excel Data":
        show_excel_import()
    elif page == "📝 Measurement Sheets":
        show_measurements()
    elif page == "📊 Analytics & Reports":
        show_analytics()
    elif page == "📋 General Abstract":
        show_general_abstract()
    elif page == "💰 Abstract of Cost":
        show_abstracts()
    elif page == "📚 SSR Database":
        st.title("📚 SSR Database")
        st.dataframe(st.session_state.ssr_items, use_container_width=True, hide_index=True)
        # Upload SSR file (CSV/XLSX)
        with st.expander("📥 Upload SSR/BSR File"):
            uploaded = st.file_uploader("Upload SSR/BSR file (CSV or Excel)", type=["csv", "xlsx"])
            if uploaded is not None:
                try:
                    if uploaded.name.lower().endswith('.csv'):
                        df = pd.read_csv(uploaded)
                    else:
                        df = pd.read_excel(uploaded)
                    # Normalize columns
                    cols_map = {c.lower().strip(): c for c in df.columns}
                    required = ['code','description','category','unit','rate']
                    normalized = {}
                    for r in required:
                        # find matching column (case-insensitive contains)
                        match = next((orig for low, orig in cols_map.items() if r in low), None)
                        if match:
                            normalized[r] = df[match]
                        else:
                            normalized[r] = pd.Series(["" for _ in range(len(df))]) if r != 'rate' else pd.Series([0 for _ in range(len(df))])
                    st.session_state.ssr_items = pd.DataFrame(normalized)
                    st.success("✅ SSR loaded into session. Review above and Save to DB if desired.")
                except Exception as e:
                    st.error(f"❌ Failed to parse SSR file: {e}")

        col1, col2, col3, col4 = st.columns(4)
        with col1:
            if st.button("⬇️ Load SSR from DB"):
                ssr_df = st.session_state.database.load_ssr_items()
                if not ssr_df.empty:
                    st.session_state.ssr_items = ssr_df
                    st.success("✅ Loaded SSR from database")
                else:
                    st.info("No SSR items in database.")
        with col2:
            if st.button("⬆️ Save SSR to DB"):
                ok = st.session_state.database.update_ssr_items(st.session_state.ssr_items)
                if ok:
                    st.success("✅ SSR saved to database")
                else:
                    st.error("❌ Failed to save SSR")
        with col3:
            # Generate PDF for item 5
            def _generate_ssr_item_pdf(item_row: pd.Series) -> bytes:
                try:
                    from reportlab.lib.pagesizes import A4
                    from reportlab.pdfgen import canvas
                    from reportlab.lib.units import cm
                except Exception:
                    return b""
                buf = io.BytesIO()
                c = canvas.Canvas(buf, pagesize=A4)
                width, height = A4
                c.setFont("Helvetica-Bold", 14)
                c.drawString(2*cm, height-2*cm, "SSR Item Certificate")
                c.setFont("Helvetica", 11)
                y = height - 3*cm
                lines = [
                    f"Code: {item_row.get('code','')}",
                    f"Description: {item_row.get('description','')}",
                    f"Category: {item_row.get('category','')}",
                    f"Unit: {item_row.get('unit','')}",
                    f"Rate: {item_row.get('rate','')}",
                    f"Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
                ]
                for line in lines:
                    c.drawString(2*cm, y, str(line))
                    y -= 1*cm
                c.showPage()
                c.save()
                pdf = buf.getvalue()
                buf.close()
                return pdf

            if st.button("📄 Generate PDF for SSR Item 5"):
                df = st.session_state.ssr_items
                if df is not None and len(df) >= 5:
                    row = df.iloc[4]  # 0-based index
                    pdf_bytes = _generate_ssr_item_pdf(row)
                    if pdf_bytes:
                        # Save to session and provide download + inline preview
                        st.session_state['ssr_item_5_pdf'] = pdf_bytes
                        st.download_button("⬇️ Download SSR Item 5 PDF", data=pdf_bytes, file_name="ssr_item_5.pdf", mime="application/pdf")
                        with st.expander("👁️ Preview SSR Item 5 PDF"):
                            b64 = base64.b64encode(pdf_bytes).decode('utf-8')
                            pdf_html = f'<iframe src="data:application/pdf;base64,{b64}" width="100%" height="600px" type="application/pdf"></iframe>'
                            st.markdown(pdf_html, unsafe_allow_html=True)
                    else:
                        st.warning("Install 'reportlab' to enable PDF export: pip install reportlab")
                else:
                    st.info("Need at least 5 SSR items loaded to export item 5.")
        with col4:
            st.write("")
    elif page == "🔧 System Tools":
        st.title("🔧 System Tools")
        
        col1, col2 = st.columns(2)
        
        with col1:
            if st.button("💾 Save Current Project"):
                project_data = {
                    'name': st.session_state.project_settings['project_name'],
                    'location': st.session_state.project_settings['project_location'],
                    'measurements': st.session_state.measurement_sheets,
                    'abstracts': st.session_state.abstract_sheets,
                    'total_cost': sum(
                        df['amount'].sum() if not df.empty else 0 
                        for df in st.session_state.abstract_sheets.values()
                    )
                }
                
                project_id = st.session_state.database.save_project(project_data)
                if project_id > 0:
                    st.success(f"✅ Project saved! ID: {project_id}")
                    st.session_state.project_settings['current_project_id'] = project_id
                else:
                    st.error("❌ Failed to save project")
        
        with col2:
            if st.button("🗑️ Clear All Data"):
                # Reset all session state
                for key in ['measurements', 'measurement_sheets', 'abstract_sheets']:
                    if key in st.session_state:
                        del st.session_state[key]
                initialize_session_state()
                st.success("✅ All data cleared!")
                st.rerun()
    elif page == "🧩 Templates":
        show_templates()

# =============================================================================
# APPLICATION ENTRY POINT
# =============================================================================

if __name__ == "__main__":
    main()
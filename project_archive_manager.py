"""
Project Archive Manager
Manages historical estimates organized by project type (Buildings, Bridges, etc.)
"""

import json
import shutil
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Dict, Optional

import pandas as pd
import streamlit as st
from openpyxl import load_workbook


class ProjectArchiveManager:
    """Manages archived project estimates organized by category"""
    
    def __init__(self, archive_root: str = "project_archives"):
        self.archive_root = Path(archive_root)
        self.archive_root.mkdir(exist_ok=True)
        
        # Create category folders
        self.categories = {
            "1_BUILDINGS": "Buildings",
            "2_BRIDGES": "Bridges",
            "3_ROADS": "Roads",
            "4_WATER_SUPPLY": "Water Supply",
            "5_DRAINAGE": "Drainage",
            "6_OTHERS": "Others"
        }
        
        for folder, name in self.categories.items():
            (self.archive_root / folder).mkdir(exist_ok=True)
        
        # Initialize metadata database
        self.db_path = self.archive_root / "archive_metadata.db"
        self._init_database()
    
    def _init_database(self):
        """Initialize archive metadata database"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS archived_projects (
                id TEXT PRIMARY KEY,
                file_name TEXT NOT NULL,
                original_name TEXT NOT NULL,
                category TEXT NOT NULL,
                project_name TEXT,
                location TEXT,
                client_name TEXT,
                estimated_cost REAL,
                date_prepared TEXT,
                engineer_name TEXT,
                file_size INTEGER,
                sheet_count INTEGER,
                row_count INTEGER,
                archived_date TEXT,
                file_path TEXT,
                notes TEXT,
                tags TEXT,
                status TEXT DEFAULT 'active'
            )
        """)
        
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS archive_tags (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_id TEXT,
                tag TEXT,
                FOREIGN KEY (project_id) REFERENCES archived_projects(id)
            )
        """)
        
        conn.commit()
        conn.close()
    
    def archive_project(self, file_path: str, category: str, metadata: Dict) -> Dict:
        """Archive a project estimate file"""
        try:
            source_path = Path(file_path)
            
            # Validate file
            if not source_path.exists():
                return {"success": False, "error": "File not found"}
            
            if source_path.suffix.lower() not in ['.xls', '.xlsx']:
                return {"success": False, "error": "Only XLS/XLSX files supported"}
            
            # Generate unique ID
            project_id = f"{category}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            
            # Create destination path
            category_folder = self.archive_root / category
            dest_filename = f"{project_id}_{source_path.name}"
            dest_path = category_folder / dest_filename
            
            # Copy file
            shutil.copy2(source_path, dest_path)
            
            # Extract file metadata
            file_metadata = self._extract_file_metadata(dest_path)
            
            # Save to database
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            cursor.execute("""
                INSERT INTO archived_projects 
                (id, file_name, original_name, category, project_name, location, 
                 client_name, estimated_cost, date_prepared, engineer_name,
                 file_size, sheet_count, row_count, archived_date, file_path, notes, tags)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                project_id,
                dest_filename,
                source_path.name,
                category,
                metadata.get('project_name', ''),
                metadata.get('location', ''),
                metadata.get('client_name', ''),
                metadata.get('estimated_cost', 0),
                metadata.get('date_prepared', ''),
                metadata.get('engineer_name', ''),
                file_metadata['file_size'],
                file_metadata['sheet_count'],
                file_metadata['row_count'],
                datetime.now().isoformat(),
                str(dest_path),
                metadata.get('notes', ''),
                json.dumps(metadata.get('tags', []))
            ))
            
            conn.commit()
            conn.close()
            
            return {
                "success": True,
                "project_id": project_id,
                "file_path": str(dest_path)
            }
            
        except Exception as e:
            return {"success": False, "error": str(e)}
    
    def _extract_file_metadata(self, file_path: Path) -> Dict:
        """Extract metadata from Excel file"""
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            
            sheet_count = len(wb.sheetnames)
            row_count = 0
            
            for sheet in wb.worksheets:
                row_count += sheet.max_row
            
            wb.close()
            
            return {
                'file_size': file_path.stat().st_size,
                'sheet_count': sheet_count,
                'row_count': row_count
            }
        except:
            return {
                'file_size': file_path.stat().st_size,
                'sheet_count': 0,
                'row_count': 0
            }
    
    def get_archived_projects(self, category: Optional[str] = None, 
                             search_term: Optional[str] = None) -> pd.DataFrame:
        """Get list of archived projects"""
        conn = sqlite3.connect(self.db_path)
        
        query = "SELECT * FROM archived_projects WHERE status = 'active'"
        params = []
        
        if category:
            query += " AND category = ?"
            params.append(category)
        
        if search_term:
            query += " AND (project_name LIKE ? OR location LIKE ? OR client_name LIKE ?)"
            search_pattern = f"%{search_term}%"
            params.extend([search_pattern, search_pattern, search_pattern])
        
        query += " ORDER BY archived_date DESC"
        
        df = pd.read_sql_query(query, conn, params=params)
        conn.close()
        
        return df
    
    def get_project_details(self, project_id: str) -> Optional[Dict]:
        """Get detailed information about an archived project"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT * FROM archived_projects WHERE id = ?
        """, (project_id,))
        
        row = cursor.fetchone()
        conn.close()
        
        if row:
            columns = [desc[0] for desc in cursor.description]
            return dict(zip(columns, row))
        return None
    
    def get_statistics(self) -> Dict:
        """Get archive statistics"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # Total projects by category
        cursor.execute("""
            SELECT category, COUNT(*) as count, SUM(estimated_cost) as total_cost
            FROM archived_projects
            WHERE status = 'active'
            GROUP BY category
        """)
        
        category_stats = {}
        for row in cursor.fetchall():
            category_stats[row[0]] = {
                'count': row[1],
                'total_cost': row[2] or 0
            }
        
        # Total statistics
        cursor.execute("""
            SELECT 
                COUNT(*) as total_projects,
                SUM(estimated_cost) as total_cost,
                SUM(file_size) as total_size,
                AVG(estimated_cost) as avg_cost
            FROM archived_projects
            WHERE status = 'active'
        """)
        
        total_stats = cursor.fetchone()
        
        conn.close()
        
        return {
            'category_stats': category_stats,
            'total_projects': total_stats[0],
            'total_cost': total_stats[1] or 0,
            'total_size': total_stats[2] or 0,
            'avg_cost': total_stats[3] or 0
        }
    
    def bulk_import(self, folder_path: str, category: str) -> Dict:
        """Bulk import multiple Excel files from a folder"""
        results = {
            'success': [],
            'failed': [],
            'total': 0
        }
        
        folder = Path(folder_path)
        if not folder.exists():
            return results
        
        excel_files = list(folder.glob('*.xls')) + list(folder.glob('*.xlsx'))
        results['total'] = len(excel_files)
        
        for file_path in excel_files:
            # Extract basic metadata from filename
            metadata = {
                'project_name': file_path.stem,
                'date_prepared': datetime.now().strftime('%Y-%m-%d'),
                'tags': [category]
            }
            
            result = self.archive_project(str(file_path), category, metadata)
            
            if result['success']:
                results['success'].append(file_path.name)
            else:
                results['failed'].append({
                    'file': file_path.name,
                    'error': result.get('error', 'Unknown error')
                })
        
        return results
    
    def export_project_list(self, category: Optional[str] = None) -> pd.DataFrame:
        """Export project list to Excel"""
        df = self.get_archived_projects(category)
        
        # Select relevant columns
        export_columns = [
            'id', 'project_name', 'location', 'client_name', 
            'estimated_cost', 'date_prepared', 'engineer_name',
            'category', 'archived_date', 'notes'
        ]
        
        return df[export_columns] if not df.empty else pd.DataFrame()


def render_archive_ui():
    """Render the archive management UI in Streamlit"""
    st.title("📚 Project Archive Manager")
    
    # Initialize archive manager
    if 'archive_manager' not in st.session_state:
        st.session_state.archive_manager = ProjectArchiveManager()
    
    archive_mgr = st.session_state.archive_manager
    
    # Tabs
    tab1, tab2, tab3, tab4 = st.tabs([
        "📊 Dashboard", 
        "📥 Import Projects", 
        "🔍 Browse Archives",
        "📤 Export"
    ])
    
    # Tab 1: Dashboard
    with tab1:
        st.subheader("Archive Statistics")
        
        stats = archive_mgr.get_statistics()
        
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric("Total Projects", stats['total_projects'])
        
        with col2:
            st.metric("Total Cost", f"₹{stats['total_cost']:,.0f}")
        
        with col3:
            st.metric("Average Cost", f"₹{stats['avg_cost']:,.0f}")
        
        with col4:
            size_mb = stats['total_size'] / (1024 * 1024)
            st.metric("Archive Size", f"{size_mb:.1f} MB")
        
        # Category breakdown
        st.subheader("Projects by Category")
        
        category_data = []
        for cat_key, cat_name in archive_mgr.categories.items():
            if cat_key in stats['category_stats']:
                cat_stats = stats['category_stats'][cat_key]
                category_data.append({
                    'Category': cat_name,
                    'Projects': cat_stats['count'],
                    'Total Cost': cat_stats['total_cost']
                })
        
        if category_data:
            df_stats = pd.DataFrame(category_data)
            
            col1, col2 = st.columns(2)
            
            with col1:
                st.dataframe(df_stats, use_container_width=True)
            
            with col2:
                import plotly.express as px
                fig = px.pie(df_stats, values='Projects', names='Category', 
                           title='Projects by Category')
                st.plotly_chart(fig, use_container_width=True)
    
    # Tab 2: Import Projects
    with tab2:
        st.subheader("Import Project Estimates")
        
        import_method = st.radio(
            "Import Method",
            ["Single File", "Bulk Import from Folder"]
        )
        
        if import_method == "Single File":
            # Single file import
            col1, col2 = st.columns([2, 1])
            
            with col1:
                uploaded_file = st.file_uploader(
                    "Upload Excel File (XLS/XLSX)",
                    type=['xls', 'xlsx']
                )
            
            with col2:
                category = st.selectbox(
                    "Category",
                    options=list(archive_mgr.categories.keys()),
                    format_func=lambda x: archive_mgr.categories[x]
                )
            
            if uploaded_file:
                with st.form("project_metadata"):
                    st.write("**Project Details**")
                    
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        project_name = st.text_input("Project Name", uploaded_file.name.rsplit('.', 1)[0])
                        location = st.text_input("Location")
                        client_name = st.text_input("Client Name")
                    
                    with col2:
                        estimated_cost = st.number_input("Estimated Cost (₹)", min_value=0.0, step=1000.0)
                        date_prepared = st.date_input("Date Prepared")
                        engineer_name = st.text_input("Engineer Name")
                    
                    notes = st.text_area("Notes")
                    tags = st.text_input("Tags (comma-separated)")
                    
                    submitted = st.form_submit_button("Archive Project")
                    
                    if submitted:
                        # Save uploaded file temporarily
                        temp_path = Path("temp") / uploaded_file.name
                        temp_path.parent.mkdir(exist_ok=True)
                        
                        with open(temp_path, "wb") as f:
                            f.write(uploaded_file.getbuffer())
                        
                        # Archive the project
                        metadata = {
                            'project_name': project_name,
                            'location': location,
                            'client_name': client_name,
                            'estimated_cost': estimated_cost,
                            'date_prepared': date_prepared.isoformat(),
                            'engineer_name': engineer_name,
                            'notes': notes,
                            'tags': [t.strip() for t in tags.split(',') if t.strip()]
                        }
                        
                        result = archive_mgr.archive_project(str(temp_path), category, metadata)
                        
                        # Clean up temp file
                        temp_path.unlink()
                        
                        if result['success']:
                            st.success(f"✅ Project archived successfully! ID: {result['project_id']}")
                        else:
                            st.error(f"❌ Error: {result['error']}")
        
        else:
            # Bulk import
            st.info("📁 Select a folder containing multiple Excel files to import")
            
            folder_path = st.text_input("Folder Path", placeholder="C:/Projects/Buildings")
            category = st.selectbox(
                "Category for all files",
                options=list(archive_mgr.categories.keys()),
                format_func=lambda x: archive_mgr.categories[x]
            )
            
            if st.button("Start Bulk Import"):
                if folder_path:
                    with st.spinner("Importing files..."):
                        results = archive_mgr.bulk_import(folder_path, category)
                    
                    st.success(f"✅ Imported {len(results['success'])} of {results['total']} files")
                    
                    if results['success']:
                        st.write("**Successfully imported:**")
                        for file in results['success']:
                            st.write(f"- {file}")
                    
                    if results['failed']:
                        st.error("**Failed imports:**")
                        for item in results['failed']:
                            st.write(f"- {item['file']}: {item['error']}")
    
    # Tab 3: Browse Archives
    with tab3:
        st.subheader("Browse Archived Projects")
        
        col1, col2 = st.columns([1, 2])
        
        with col1:
            filter_category = st.selectbox(
                "Filter by Category",
                options=["All"] + list(archive_mgr.categories.keys()),
                format_func=lambda x: "All Categories" if x == "All" else archive_mgr.categories.get(x, x)
            )
        
        with col2:
            search_term = st.text_input("🔍 Search", placeholder="Search by project name, location, or client")
        
        # Get projects
        category_filter = None if filter_category == "All" else filter_category
        df = archive_mgr.get_archived_projects(category_filter, search_term if search_term else None)
        
        if not df.empty:
            st.write(f"**Found {len(df)} projects**")
            
            # Display projects
            for idx, row in df.iterrows():
                with st.expander(f"📄 {row['project_name']} - {row['location']}"):
                    col1, col2, col3 = st.columns(3)
                    
                    with col1:
                        st.write(f"**Category:** {archive_mgr.categories.get(row['category'], row['category'])}")
                        st.write(f"**Client:** {row['client_name']}")
                        st.write(f"**Location:** {row['location']}")
                    
                    with col2:
                        st.write(f"**Cost:** ₹{row['estimated_cost']:,.0f}")
                        st.write(f"**Date Prepared:** {row['date_prepared']}")
                        st.write(f"**Engineer:** {row['engineer_name']}")
                    
                    with col3:
                        st.write(f"**Sheets:** {row['sheet_count']}")
                        st.write(f"**Rows:** {row['row_count']}")
                        st.write(f"**Archived:** {row['archived_date'][:10]}")
                    
                    if row['notes']:
                        st.write(f"**Notes:** {row['notes']}")
                    
                    # Action buttons
                    col1, col2, col3 = st.columns(3)
                    
                    with col1:
                        if st.button("📂 Open File", key=f"open_{row['id']}"):
                            st.info(f"File location: {row['file_path']}")
                    
                    with col2:
                        if st.button("📥 Import to Project", key=f"import_{row['id']}"):
                            st.info("Import functionality - integrate with main app")
                    
                    with col3:
                        if st.button("🗑️ Delete", key=f"delete_{row['id']}"):
                            st.warning("Delete functionality - to be implemented")
        else:
            st.info("No projects found. Start by importing your existing estimates!")
    
    # Tab 4: Export
    with tab4:
        st.subheader("Export Project List")
        
        export_category = st.selectbox(
            "Export Category",
            options=["All"] + list(archive_mgr.categories.keys()),
            format_func=lambda x: "All Categories" if x == "All" else archive_mgr.categories.get(x, x)
        )
        
        if st.button("Generate Export"):
            category_filter = None if export_category == "All" else export_category
            export_df = archive_mgr.export_project_list(category_filter)
            
            if not export_df.empty:
                # Convert to Excel
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    export_df.to_excel(writer, index=False, sheet_name='Projects')
                
                output.seek(0)
                
                st.download_button(
                    label="📥 Download Excel",
                    data=output,
                    file_name=f"project_archive_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
                st.success(f"✅ Export ready! {len(export_df)} projects")
            else:
                st.info("No projects to export")


if __name__ == "__main__":
    render_archive_ui()

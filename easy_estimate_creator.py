"""
Easy Estimate Creator - Simple Script for Creating Multiple Estimates
======================================================================
Just run: python easy_estimate_creator.py
"""
from copy import copy
from datetime import datetime
from pathlib import Path

import openpyxl

# ============ CONFIGURATION ============
NUM_ESTIMATES = 5
LINES_PER_ITEM = 3
ITEMS_TO_MODIFY = 5
ABSTRACT_INCREASE = 15  # Percentage

SOURCE_FOLDER = "attached_assets"
OUTPUT_FOLDER = "generated_estimates"

PROJECTS = [
    "Modern Residential",
    "Smart Library", 
    "Underground Parking",
    "Municipal Building",
    "Sports Hall"
]
# =======================================

def create_estimates():
    """Main function to create estimates"""
    
    print("\n" + "="*80)
    print(f"Creating {NUM_ESTIMATES} Estimates")
    print("="*80 + "\n")
    
    # Setup
    output = Path(OUTPUT_FOLDER)
    output.mkdir(exist_ok=True)
    
    source_files = list(Path(SOURCE_FOLDER).glob("*.xlsx"))[:NUM_ESTIMATES]
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    created = []
    
    # Process each file
    for idx, (src, proj) in enumerate(zip(source_files, PROJECTS), 1):
        print(f"\n{idx}/{NUM_ESTIMATES}: {proj}")
        print("-" * 40)
        
        try:
            # Load workbook
            wb = openpyxl.load_workbook(src)
            
            # Find measurement sheet
            mes_sheet = next((s for s in wb.sheetnames if 'MES' in s.upper()), None)
            if not mes_sheet:
                print("  ⚠️  No measurement sheet")
                continue
            
            ws = wb[mes_sheet]
            print(f"  📊 Sheet: {mes_sheet}")
            
            # Find data rows
            data_rows = []
            for row_num in range(1, ws.max_row + 1):
                row = ws[row_num]
                if any(isinstance(c.value, (int, float)) and c.value > 0 for c in row[2:8]):
                    data_rows.append(row_num)
            
            print(f"  ✅ Found {len(data_rows)} measurement rows")
            
            # Add new lines
            added = 0
            offset = 0
            
            for orig_row in data_rows[:ITEMS_TO_MODIFY]:
                insert_at = orig_row + offset + 1
                orig_cells = ws[orig_row]
                
                for var in range(1, LINES_PER_ITEM + 1):
                    ws.insert_rows(insert_at)
                    
                    for col_idx, orig_cell in enumerate(orig_cells, 1):
                        new_cell = ws.cell(insert_at, col_idx)
                        
                        # Copy style
                        if orig_cell.has_style:
                            new_cell.font = copy(orig_cell.font)
                            new_cell.border = copy(orig_cell.border)
                            new_cell.fill = copy(orig_cell.fill)
                            new_cell.number_format = copy(orig_cell.number_format)
                            new_cell.alignment = copy(orig_cell.alignment)
                        
                        # Copy and modify value
                        val = orig_cell.value
                        
                        if val is None:
                            new_cell.value = val
                        elif isinstance(val, str) and len(val) > 3 and col_idx <= 2:
                            new_cell.value = f"{val} - V{var}"
                        elif isinstance(val, (int, float)):
                            multipliers = [0.85, 1.0, 1.15]
                            new_val = val * multipliers[var-1]
                            new_cell.value = int(round(new_val)) if col_idx == 3 else round(new_val, 2)
                        else:
                            new_cell.value = val
                    
                    insert_at += 1
                    added += 1
                
                offset += LINES_PER_ITEM
            
            print(f"  ✅ Added {added} new lines")
            
            # Update abstract
            abs_sheet = next((s for s in wb.sheetnames if 'ABS' in s.upper() and 'GEN' not in s.upper()), None)
            
            if abs_sheet:
                abs_ws = wb[abs_sheet]
                qty_updated = 0
                
                for row in abs_ws.iter_rows():
                    for cell in row:
                        if isinstance(cell.value, (int, float)) and cell.value > 10:
                            cell.value = round(cell.value * (1 + ABSTRACT_INCREASE/100), 2)
                            qty_updated += 1
                
                print(f"  ✅ Abstract updated: +{ABSTRACT_INCREASE}% ({qty_updated} values)")
            
            # Save
            output_file = output / f"Est_{idx}_{proj.replace(' ', '_')}_{timestamp}.xlsx"
            wb.save(output_file)
            
            print(f"  💾 {output_file.name}")
            created.append(output_file.name)
            
        except Exception as e:
            print(f"  ❌ Error: {e}")
    
    # Summary
    print(f"\n{'='*80}")
    print(f"✅ Created {len(created)} estimates in: {output.absolute()}")
    print("="*80 + "\n")

if __name__ == "__main__":
    create_estimates()

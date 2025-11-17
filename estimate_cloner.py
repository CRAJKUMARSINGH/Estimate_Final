"""
Estimate Cloner & Modifier
Create new estimates from archived estimates with modifications
"""

from datetime import datetime
from pathlib import Path
from typing import Dict, Optional

import openpyxl
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# PDF generation imports
try:
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm, inch
    from reportlab.pdfgen import canvas
    from reportlab.platypus import (PageBreak, Paragraph, SimpleDocTemplate,
                                    Spacer, Table, TableStyle)
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

class EstimateCloner:
    """Clone and modify existing estimates"""
    
    def __init__(self):
        self.source_estimate = None
        self.modified_estimate = None
        self.modifications = []
    
    def load_estimate(self, file_path: str) -> Dict:
        """Load an archived estimate"""
        try:
            wb = load_workbook(file_path, data_only=False)
            
            estimate_data = {
                'file_path': file_path,
                'file_name': Path(file_path).name,
                'sheets': {},
                'workbook': wb
            }
            
            # Load each sheet
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                
                # Convert to DataFrame
                data = []
                for row in sheet.iter_rows(values_only=True):
                    data.append(row)
                
                if data:
                    df = pd.DataFrame(data[1:], columns=data[0])
                    estimate_data['sheets'][sheet_name] = df
            
            self.source_estimate = estimate_data
            return estimate_data
            
        except Exception as e:
            st.error(f"Error loading estimate: {e}")
            return None
    
    def modify_measurement(self, sheet_name: str, row_index: int, 
                          column: str, new_value: any) -> bool:
        """Modify a measurement value"""
        try:
            if sheet_name in self.source_estimate['sheets']:
                df = self.source_estimate['sheets'][sheet_name].copy()
                df.at[row_index, column] = new_value
                
                # Recalculate if it's a quantity/rate change
                if column in ['Quantity', 'Rate']:
                    if 'Quantity' in df.columns and 'Rate' in df.columns and 'Amount' in df.columns:
                        df['Amount'] = pd.to_numeric(df['Quantity'], errors='coerce') * \
                                      pd.to_numeric(df['Rate'], errors='coerce')
                
                self.source_estimate['sheets'][sheet_name] = df
                
                self.modifications.append({
                    'type': 'modify',
                    'sheet': sheet_name,
                    'row': row_index,
                    'column': column,
                    'value': new_value,
                    'timestamp': datetime.now().isoformat()
                })
                
                return True
        except Exception as e:
            st.error(f"Error modifying measurement: {e}")
            return False
    
    def add_bsr_item(self, sheet_name: str, item_data: Dict, 
                     position: Optional[int] = None) -> bool:
        """Add a new BSR item to the estimate"""
        try:
            if sheet_name in self.source_estimate['sheets']:
                df = self.source_estimate['sheets'][sheet_name].copy()
                
                # Create new row
                new_row = pd.DataFrame([item_data])
                
                # Insert at position or append
                if position is not None and position < len(df):
                    df = pd.concat([
                        df.iloc[:position],
                        new_row,
                        df.iloc[position:]
                    ]).reset_index(drop=True)
                else:
                    df = pd.concat([df, new_row], ignore_index=True)
                
                # Recalculate amounts
                if 'Quantity' in df.columns and 'Rate' in df.columns and 'Amount' in df.columns:
                    df['Amount'] = pd.to_numeric(df['Quantity'], errors='coerce') * \
                                  pd.to_numeric(df['Rate'], errors='coerce')
                
                self.source_estimate['sheets'][sheet_name] = df
                
                self.modifications.append({
                    'type': 'add',
                    'sheet': sheet_name,
                    'item': item_data,
                    'position': position,
                    'timestamp': datetime.now().isoformat()
                })
                
                return True
        except Exception as e:
            st.error(f"Error adding BSR item: {e}")
            return False
    
    def delete_item(self, sheet_name: str, row_index: int) -> bool:
        """Delete an item from the estimate"""
        try:
            if sheet_name in self.source_estimate['sheets']:
                df = self.source_estimate['sheets'][sheet_name].copy()
                df = df.drop(row_index).reset_index(drop=True)
                self.source_estimate['sheets'][sheet_name] = df
                
                self.modifications.append({
                    'type': 'delete',
                    'sheet': sheet_name,
                    'row': row_index,
                    'timestamp': datetime.now().isoformat()
                })
                
                return True
        except Exception as e:
            st.error(f"Error deleting item: {e}")
            return False
    
    def recalculate_totals(self, sheet_name: str) -> bool:
        """Recalculate all totals in a sheet"""
        try:
            if sheet_name in self.source_estimate['sheets']:
                df = self.source_estimate['sheets'][sheet_name].copy()
                
                # Recalculate amounts
                if 'Quantity' in df.columns and 'Rate' in df.columns:
                    if 'Amount' not in df.columns:
                        df['Amount'] = 0
                    
                    df['Amount'] = pd.to_numeric(df['Quantity'], errors='coerce').fillna(0) * \
                                  pd.to_numeric(df['Rate'], errors='coerce').fillna(0)
                
                self.source_estimate['sheets'][sheet_name] = df
                return True
        except Exception as e:
            st.error(f"Error recalculating totals: {e}")
            return False
    
    def save_as_new_estimate(self, output_path: str, project_info: Dict) -> bool:
        """Save modified estimate as new file"""
        try:
            # Create new workbook
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            # Add each sheet
            for sheet_name, df in self.source_estimate['sheets'].items():
                ws = wb.create_sheet(sheet_name)
                
                # Write data
                for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                    for c_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=r_idx, column=c_idx, value=value)
                        
                        # Style header row
                        if r_idx == 1:
                            cell.font = Font(bold=True, color="FFFFFF")
                            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # Add metadata sheet
            meta_ws = wb.create_sheet("Metadata", 0)
            meta_ws['A1'] = "Project Information"
            meta_ws['A1'].font = Font(bold=True, size=14)
            
            row = 3
            for key, value in project_info.items():
                meta_ws[f'A{row}'] = key
                meta_ws[f'B{row}'] = value
                meta_ws[f'A{row}'].font = Font(bold=True)
                row += 1
            
            # Add modifications log
            row += 2
            meta_ws[f'A{row}'] = "Modifications Log"
            meta_ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            
            for mod in self.modifications:
                meta_ws[f'A{row}'] = f"{mod['type'].upper()}: {mod.get('sheet', '')} - {mod.get('timestamp', '')}"
                row += 1
            
            # Save
            wb.save(output_path)
            
            return True
            
        except Exception as e:
            st.error(f"Error saving estimate: {e}")
            return False


def render_estimate_cloner_ui():
    """Render the Estimate Cloner UI"""
    st.title("🔄 Estimate Cloner & Modifier")
    
    st.markdown("""
    Create new estimates from archived estimates with modifications:
    - Load existing estimate
    - Modify measurements
    - Add new BSR items
    - Update calculations
    - Save as new estimate
    """)
    
    # Initialize cloner
    if 'cloner' not in st.session_state:
        st.session_state.cloner = EstimateCloner()
    
    cloner = st.session_state.cloner
    
    # Tabs
    tab1, tab2, tab3, tab4 = st.tabs([
        "📂 Load Estimate",
        "✏️ Modify",
        "➕ Add Items",
        "💾 Save New"
    ])
    
    # Tab 1: Load Estimate
    with tab1:
        st.subheader("Load Archived Estimate")
        
        # Browse archived estimates
        archive_path = Path("project_archives")
        
        if archive_path.exists():
            # Get all Excel files
            excel_files = []
            for category in ['1_BUILDINGS', '2_BRIDGES', '3_ROADS', '4_WATER_SUPPLY', '5_DRAINAGE', '6_OTHERS']:
                cat_path = archive_path / category
                if cat_path.exists():
                    files = list(cat_path.glob("*.xlsx"))
                    excel_files.extend([(f, category) for f in files])
            
            if excel_files:
                # Create selection
                file_options = {f"{cat}: {f.name}": (f, cat) for f, cat in excel_files}
                
                selected = st.selectbox(
                    "Select Estimate to Clone",
                    options=list(file_options.keys())
                )
                
                if st.button("📂 Load Estimate"):
                    file_path, category = file_options[selected]
                    
                    with st.spinner("Loading estimate..."):
                        estimate_data = cloner.load_estimate(str(file_path))
                    
                    if estimate_data:
                        st.success(f"✅ Loaded: {estimate_data['file_name']}")
                        st.session_state.loaded_estimate = estimate_data
                        
                        # Show summary
                        st.write("**Sheets:**")
                        for sheet_name, df in estimate_data['sheets'].items():
                            st.write(f"- {sheet_name}: {len(df)} rows")
            else:
                st.info("No archived estimates found. Add estimates to project_archives first.")
        
        # Show loaded estimate
        if hasattr(st.session_state, 'loaded_estimate'):
            st.markdown("---")
            st.subheader("Loaded Estimate Preview")
            
            estimate = st.session_state.loaded_estimate
            
            sheet_to_view = st.selectbox(
                "Select Sheet to Preview",
                options=list(estimate['sheets'].keys())
            )
            
            if sheet_to_view:
                df = estimate['sheets'][sheet_to_view]
                st.dataframe(df.head(20), use_container_width=True)
                st.info(f"Showing first 20 of {len(df)} rows")
    
    # Tab 2: Modify
    with tab2:
        st.subheader("Modify Measurements")
        
        if not hasattr(st.session_state, 'loaded_estimate'):
            st.warning("⚠️ Please load an estimate first (Tab 1)")
        else:
            estimate = st.session_state.loaded_estimate
            
            # Select sheet
            sheet_name = st.selectbox(
                "Select Sheet to Modify",
                options=list(estimate['sheets'].keys()),
                key="modify_sheet"
            )
            
            if sheet_name:
                df = estimate['sheets'][sheet_name]
                
                st.write(f"**Total rows:** {len(df)}")
                
                # Show editable dataframe
                st.write("**Edit values below:**")
                
                edited_df = st.data_editor(
                    df,
                    use_container_width=True,
                    num_rows="dynamic",
                    key=f"editor_{sheet_name}"
                )
                
                col1, col2 = st.columns(2)
                
                with col1:
                    if st.button("💾 Apply Changes"):
                        cloner.source_estimate['sheets'][sheet_name] = edited_df
                        cloner.recalculate_totals(sheet_name)
                        st.success("✅ Changes applied and totals recalculated!")
                        st.rerun()
                
                with col2:
                    if st.button("🔄 Recalculate Totals"):
                        cloner.recalculate_totals(sheet_name)
                        st.success("✅ Totals recalculated!")
                        st.rerun()
    
    # Tab 3: Add Items
    with tab3:
        st.subheader("Add New BSR Items")
        
        if not hasattr(st.session_state, 'loaded_estimate'):
            st.warning("⚠️ Please load an estimate first (Tab 1)")
        else:
            estimate = st.session_state.loaded_estimate
            
            # Select sheet
            sheet_name = st.selectbox(
                "Select Sheet to Add Item",
                options=list(estimate['sheets'].keys()),
                key="add_sheet"
            )
            
            if sheet_name:
                df = estimate['sheets'][sheet_name]
                columns = df.columns.tolist()
                
                st.write("**Add New Item:**")
                
                with st.form("add_item_form"):
                    new_item = {}
                    
                    cols = st.columns(3)
                    for i, col_name in enumerate(columns):
                        with cols[i % 3]:
                            new_item[col_name] = st.text_input(col_name)
                    
                    position = st.number_input(
                        "Insert at position (0 = start, -1 = end)",
                        min_value=-1,
                        max_value=len(df),
                        value=-1
                    )
                    
                    submitted = st.form_submit_button("➕ Add Item")
                    
                    if submitted:
                        pos = None if position == -1 else position
                        if cloner.add_bsr_item(sheet_name, new_item, pos):
                            st.success("✅ Item added successfully!")
                            st.rerun()
    
    # Tab 4: Save New
    with tab4:
        st.subheader("Save as New Estimate")
        
        if not hasattr(st.session_state, 'loaded_estimate'):
            st.warning("⚠️ Please load an estimate first (Tab 1)")
        else:
            st.write("**Project Information:**")
            
            with st.form("save_estimate_form"):
                col1, col2 = st.columns(2)
                
                with col1:
                    project_name = st.text_input("Project Name*", value="New Project")
                    location = st.text_input("Location*")
                    client_name = st.text_input("Client Name")
                
                with col2:
                    engineer_name = st.text_input("Engineer Name")
                    date_prepared = st.date_input("Date Prepared")
                    estimated_cost = st.number_input("Estimated Cost (₹)", min_value=0.0)
                
                notes = st.text_area("Notes")
                
                # Output location
                output_category = st.selectbox(
                    "Save to Category",
                    options=['1_BUILDINGS', '2_BRIDGES', '3_ROADS', '4_WATER_SUPPLY', '5_DRAINAGE', '6_OTHERS']
                )
                
                output_filename = st.text_input(
                    "Output Filename",
                    value=f"Modified_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                )
                
                submitted = st.form_submit_button("💾 Save New Estimate")
                
                if submitted:
                    if not project_name or not location:
                        st.error("❌ Project Name and Location are required!")
                    else:
                        # Prepare project info
                        project_info = {
                            'Project Name': project_name,
                            'Location': location,
                            'Client Name': client_name,
                            'Engineer Name': engineer_name,
                            'Date Prepared': date_prepared.isoformat(),
                            'Estimated Cost': estimated_cost,
                            'Notes': notes,
                            'Created From': st.session_state.loaded_estimate['file_name'],
                            'Created Date': datetime.now().isoformat(),
                            'Modifications': len(cloner.modifications)
                        }
                        
                        # Output path
                        output_path = Path("project_archives") / output_category / output_filename
                        output_path.parent.mkdir(parents=True, exist_ok=True)
                        
                        with st.spinner("Saving new estimate..."):
                            if cloner.save_as_new_estimate(str(output_path), project_info):
                                st.success(f"✅ New estimate saved: {output_path}")
                                st.balloons()
                                
                                # Show summary
                                st.write("**Summary:**")
                                st.write(f"- Original: {st.session_state.loaded_estimate['file_name']}")
                                st.write(f"- New: {output_filename}")
                                st.write(f"- Modifications: {len(cloner.modifications)}")
                                st.write(f"- Location: {output_path}")
                                
                                # PDF Export
                                st.markdown("---")
                                st.write("**Export to PDF:**")
                                
                                col1, col2 = st.columns(2)
                                
                                with col1:
                                    if st.button("📄 Generate PDF (A4)"):
                                        pdf_filename = output_filename.replace('.xlsx', '.pdf')
                                        pdf_path = output_path.parent / pdf_filename
                                        
                                        with st.spinner("Generating PDF..."):
                                            if cloner.export_to_pdf(str(pdf_path), project_info):
                                                st.success(f"✅ PDF generated: {pdf_filename}")
                                                
                                                # Provide download
                                                with open(pdf_path, 'rb') as f:
                                                    pdf_bytes = f.read()
                                                
                                                st.download_button(
                                                    "📥 Download PDF",
                                                    data=pdf_bytes,
                                                    file_name=pdf_filename,
                                                    mime="application/pdf"
                                                )
                                
                                with col2:
                                    # Reset
                                    if st.button("🔄 Start New Clone"):
                                        del st.session_state.loaded_estimate
                                        st.session_state.cloner = EstimateCloner()
                                        st.rerun()


    def export_to_pdf(self, output_path: str, project_info: Dict) -> bool:
        """Export estimate to professional A4 PDF"""
        if not REPORTLAB_AVAILABLE:
            st.error("ReportLab not installed. Run: pip install reportlab")
            return False
        
        try:
            # Create PDF document
            doc = SimpleDocTemplate(
                output_path,
                pagesize=A4,
                rightMargin=1.5*cm,
                leftMargin=1.5*cm,
                topMargin=2*cm,
                bottomMargin=2*cm
            )
            
            # Container for PDF elements
            elements = []
            
            # Styles
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                textColor=colors.HexColor('#1f4e79'),
                spaceAfter=30,
                alignment=TA_CENTER,
                fontName='Helvetica-Bold'
            )
            
            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontSize=12,
                textColor=colors.HexColor('#2d5aa0'),
                spaceAfter=12,
                spaceBefore=12,
                fontName='Helvetica-Bold'
            )
            
            # Title
            elements.append(Paragraph("CONSTRUCTION ESTIMATE", title_style))
            elements.append(Spacer(1, 0.3*cm))
            
            # Project Information Box
            project_data = [
                ['Project Name:', project_info.get('Project Name', 'N/A')],
                ['Location:', project_info.get('Location', 'N/A')],
                ['Client:', project_info.get('Client Name', 'N/A')],
                ['Engineer:', project_info.get('Engineer Name', 'N/A')],
                ['Date:', project_info.get('Date Prepared', 'N/A')],
                ['Estimated Cost:', f"₹ {project_info.get('Estimated Cost', 0):,.2f}"]
            ]
            
            project_table = Table(project_data, colWidths=[4*cm, 12*cm])
            project_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e7e6e6')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
                ('ALIGN', (1, 0), (1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 8),
                ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            
            elements.append(project_table)
            elements.append(Spacer(1, 0.5*cm))
            
            # Add each sheet
            for sheet_name, df in self.source_estimate['sheets'].items():
                # Skip metadata sheet
                if sheet_name.lower() == 'metadata':
                    continue
                
                # Sheet heading
                elements.append(PageBreak())
                elements.append(Paragraph(f"{sheet_name}", heading_style))
                elements.append(Spacer(1, 0.3*cm))
                
                # Prepare data - select key columns
                key_columns = []
                for col in ['Sr No', 'Description', 'Quantity', 'Unit', 'Rate', 'Amount']:
                    if col in df.columns:
                        key_columns.append(col)
                
                # If no standard columns, use first 6
                if not key_columns:
                    key_columns = df.columns.tolist()[:6]
                
                df_display = df[key_columns].copy()
                
                # Convert to list
                table_data = [key_columns]
                
                # Add rows (limit for PDF size)
                max_rows = 40
                for idx, row in df_display.head(max_rows).iterrows():
                    row_data = []
                    for val in row:
                        # Format numbers
                        if isinstance(val, (int, float)):
                            if val > 1000:
                                row_data.append(f'{val:,.2f}')
                            else:
                                row_data.append(f'{val:.2f}')
                        else:
                            row_data.append(str(val)[:40] if val else '')
                    table_data.append(row_data)
                
                if len(df) > max_rows:
                    table_data.append(['...' for _ in range(len(key_columns))])
                
                # Calculate column widths
                available_width = 17*cm
                col_widths = [available_width / len(key_columns)] * len(key_columns)
                
                # Adjust for description column
                if 'Description' in key_columns:
                    desc_idx = key_columns.index('Description')
                    col_widths[desc_idx] = 6*cm
                    remaining = (available_width - 6*cm) / (len(key_columns) - 1)
                    for i in range(len(col_widths)):
                        if i != desc_idx:
                            col_widths[i] = remaining
                
                # Create table
                data_table = Table(table_data, colWidths=col_widths, repeatRows=1)
                data_table.setStyle(TableStyle([
                    # Header
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 8),
                    
                    # Data
                    ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                    ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # Sr No
                    ('ALIGN', (1, 1), (1, -1), 'LEFT'),    # Description
                    ('ALIGN', (2, 1), (-1, -1), 'RIGHT'),  # Numbers
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 7),
                    ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('LEFTPADDING', (0, 0), (-1, -1), 3),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                    ('TOPPADDING', (0, 0), (-1, -1), 2),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
                ]))
                
                elements.append(data_table)
                elements.append(Spacer(1, 0.4*cm))
                
                # Add sheet total
                if 'Amount' in df.columns:
                    try:
                        total = pd.to_numeric(df['Amount'], errors='coerce').sum()
                        summary_data = [[f'{sheet_name} Total:', f'₹ {total:,.2f}']]
                        summary_table = Table(summary_data, colWidths=[10*cm, 6*cm])
                        summary_table.setStyle(TableStyle([
                            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#d9e1f2')),
                            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
                            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                            ('FONTSIZE', (0, 0), (-1, -1), 10),
                            ('BOX', (0, 0), (-1, -1), 1, colors.black),
                            ('LEFTPADDING', (0, 0), (-1, -1), 8),
                            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                            ('TOPPADDING', (0, 0), (-1, -1), 6),
                            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                        ]))
                        elements.append(summary_table)
                    except:
                        pass
            
            # Build PDF
            doc.build(elements)
            
            return True
            
        except Exception as e:
            st.error(f"Error generating PDF: {e}")
            return False


if __name__ == "__main__":
    render_estimate_cloner_ui()

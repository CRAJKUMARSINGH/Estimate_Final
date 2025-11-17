"""
Dynamic Template Renderer
Auto-detects input/output cells from Excel templates and generates UI forms
"""
from dataclasses import dataclass
from typing import Any, Dict, List, Optional

import openpyxl


@dataclass
class TemplateField:
    """Represents a field in a template"""
    cell_ref: str
    sheet_name: str
    field_type: str  # 'input' or 'output'
    label: str
    value: Any
    data_type: str  # 'number', 'text', 'formula', 'date'
    validation: Optional[Dict] = None
    color: Optional[str] = None
    formula: Optional[str] = None


class DynamicTemplateRenderer:
    """Renders Excel templates as dynamic UI forms"""
    
    # Color codes for input/output detection
    INPUT_COLORS = ['FFFFFF00', 'FFFF00', 'FFFFFF99']  # Yellow shades
    OUTPUT_COLORS = ['FF90EE90', '90EE90', 'FFCCFFCC', 'FF00FF00']  # Green shades
    
    def __init__(self):
        self.input_fields: List[TemplateField] = []
        self.output_fields: List[TemplateField] = []
        self.formulas: Dict[str, str] = {}
        self.named_ranges: Dict[str, str] = {}
    
    def analyze_template(self, file_path: str) -> Dict[str, Any]:
        """
        Analyze Excel template and extract input/output fields
        
        Args:
            file_path: Path to Excel template file
            
        Returns:
            Dictionary containing template structure
        """
        try:
            wb = openpyxl.load_workbook(file_path, data_only=False)
            
            # Extract named ranges
            if hasattr(wb, 'defined_names'):
                for named_range in wb.defined_names.definedName:
                    self.named_ranges[named_range.name] = str(named_range.value)
            
            # Analyze each sheet
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                self._analyze_sheet(sheet, sheet_name)
            
            wb.close()
            
            return {
                'input_fields': [self._field_to_dict(f) for f in self.input_fields],
                'output_fields': [self._field_to_dict(f) for f in self.output_fields],
                'formulas': self.formulas,
                'named_ranges': self.named_ranges,
                'total_inputs': len(self.input_fields),
                'total_outputs': len(self.output_fields)
            }
            
        except Exception as e:
            return {'error': str(e)}
    
    def _analyze_sheet(self, sheet, sheet_name: str):
        """Analyze individual sheet for input/output cells"""
        
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                
                # Check cell color
                field_type = self._detect_field_type(cell)
                
                if field_type:
                    # Extract field information
                    field = self._create_field(cell, sheet_name, field_type)
                    
                    if field_type == 'input':
                        self.input_fields.append(field)
                    else:
                        self.output_fields.append(field)
                    
                    # Store formula if present
                    if cell.data_type == 'f':
                        cell_ref = f"{sheet_name}!{cell.coordinate}"
                        self.formulas[cell_ref] = str(cell.value)
    
    def _detect_field_type(self, cell) -> Optional[str]:
        """Detect if cell is input or output based on color"""
        
        if not cell.fill or not cell.fill.start_color:
            return None
        
        color = cell.fill.start_color.rgb
        
        if not color or color == '00000000':
            return None
        
        # Check for input colors (yellow)
        if any(c in color for c in self.INPUT_COLORS):
            return 'input'
        
        # Check for output colors (green)
        if any(c in color for c in self.OUTPUT_COLORS):
            return 'output'
        
        return None
    
    def _create_field(self, cell, sheet_name: str, field_type: str) -> TemplateField:
        """Create TemplateField from cell"""
        
        # Try to extract label from adjacent cell (left or above)
        label = self._extract_label(cell)
        
        # Determine data type
        data_type = self._determine_data_type(cell)
        
        # Extract validation rules
        validation = self._extract_validation(cell)
        
        # Get color
        color = cell.fill.start_color.rgb if cell.fill and cell.fill.start_color else None
        
        # Get formula if present
        formula = str(cell.value) if cell.data_type == 'f' else None
        
        return TemplateField(
            cell_ref=cell.coordinate,
            sheet_name=sheet_name,
            field_type=field_type,
            label=label,
            value=cell.value,
            data_type=data_type,
            validation=validation,
            color=color,
            formula=formula
        )
    
    def _extract_label(self, cell) -> str:
        """Try to extract label from adjacent cells"""
        
        sheet = cell.parent
        row = cell.row
        col = cell.column
        
        # Check left cell
        if col > 1:
            left_cell = sheet.cell(row=row, column=col-1)
            if left_cell.value and isinstance(left_cell.value, str):
                return left_cell.value.strip()
        
        # Check above cell
        if row > 1:
            above_cell = sheet.cell(row=row-1, column=col)
            if above_cell.value and isinstance(above_cell.value, str):
                return above_cell.value.strip()
        
        # Use cell reference as fallback
        return f"Field {cell.coordinate}"
    
    def _determine_data_type(self, cell) -> str:
        """Determine data type of cell"""
        
        if cell.data_type == 'f':
            return 'formula'
        elif cell.data_type == 'n':
            return 'number'
        elif cell.data_type == 'd':
            return 'date'
        elif cell.data_type == 'b':
            return 'boolean'
        else:
            return 'text'
    
    def _extract_validation(self, cell) -> Optional[Dict]:
        """Extract data validation rules from cell"""
        
        if not hasattr(cell, 'data_validation') or not cell.data_validation:
            return None
        
        dv = cell.data_validation
        
        validation = {}
        
        if dv.type:
            validation['type'] = dv.type
        
        if dv.formula1:
            validation['min'] = dv.formula1
        
        if dv.formula2:
            validation['max'] = dv.formula2
        
        if dv.allow_blank is not None:
            validation['allow_blank'] = dv.allow_blank
        
        return validation if validation else None
    
    def _field_to_dict(self, field: TemplateField) -> Dict:
        """Convert TemplateField to dictionary"""
        return {
            'cell_ref': field.cell_ref,
            'sheet_name': field.sheet_name,
            'field_type': field.field_type,
            'label': field.label,
            'value': field.value,
            'data_type': field.data_type,
            'validation': field.validation,
            'color': field.color,
            'formula': field.formula
        }
    
    def generate_ui_config(self) -> Dict[str, Any]:
        """Generate UI configuration for rendering forms"""
        
        # Group input fields by sheet
        inputs_by_sheet = {}
        for field in self.input_fields:
            if field.sheet_name not in inputs_by_sheet:
                inputs_by_sheet[field.sheet_name] = []
            inputs_by_sheet[field.sheet_name].append(field)
        
        # Group output fields by sheet
        outputs_by_sheet = {}
        for field in self.output_fields:
            if field.sheet_name not in outputs_by_sheet:
                outputs_by_sheet[field.sheet_name] = []
            outputs_by_sheet[field.sheet_name].append(field)
        
        return {
            'inputs_by_sheet': {
                sheet: [self._field_to_dict(f) for f in fields]
                for sheet, fields in inputs_by_sheet.items()
            },
            'outputs_by_sheet': {
                sheet: [self._field_to_dict(f) for f in fields]
                for sheet, fields in outputs_by_sheet.items()
            },
            'formulas': self.formulas,
            'named_ranges': self.named_ranges
        }
    
    def update_template(self, file_path: str, input_values: Dict[str, Any], output_path: str):
        """
        Update template with input values and save
        
        Args:
            file_path: Original template path
            input_values: Dictionary of cell_ref: value pairs
            output_path: Path to save updated file
        """
        wb = openpyxl.load_workbook(file_path)
        
        # Update input cells
        for cell_ref, value in input_values.items():
            # Parse cell reference (Sheet!A1 format)
            if '!' in cell_ref:
                sheet_name, cell_coord = cell_ref.split('!')
                sheet = wb[sheet_name]
                sheet[cell_coord] = value
        
        # Save updated workbook
        wb.save(output_path)
        wb.close()
        
        return output_path

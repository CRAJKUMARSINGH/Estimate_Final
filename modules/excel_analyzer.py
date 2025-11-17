"""
Excel Structure Analyzer Module
Analyzes Excel files to help debug import issues and understand structure
"""
from pathlib import Path
from typing import Any, Dict, List

import openpyxl


class ExcelAnalyzer:
    """Analyzes Excel file structure for debugging and validation"""
    
    def __init__(self):
        self.analysis_results = {}
    
    def analyze_file(self, file_path: str) -> Dict[str, Any]:
        """
        Perform comprehensive analysis of Excel file
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            Dictionary containing complete analysis
        """
        try:
            wb = openpyxl.load_workbook(file_path, data_only=False)
            
            analysis = {
                'file_name': Path(file_path).name,
                'file_size': Path(file_path).stat().st_size,
                'total_sheets': len(wb.sheetnames),
                'sheet_names': wb.sheetnames,
                'sheets': {},
                'has_formulas': False,
                'has_named_ranges': False,
                'has_data_validation': False,
                'summary': {}
            }
            
            # Check for named ranges
            if hasattr(wb, 'defined_names') and wb.defined_names.definedName:
                analysis['has_named_ranges'] = True
                analysis['named_ranges'] = [nr.name for nr in wb.defined_names.definedName]
            
            # Analyze each sheet
            for sheet_name in wb.sheetnames:
                sheet_analysis = self._analyze_sheet(wb[sheet_name], sheet_name)
                analysis['sheets'][sheet_name] = sheet_analysis
                
                if sheet_analysis['has_formulas']:
                    analysis['has_formulas'] = True
                if sheet_analysis['has_data_validation']:
                    analysis['has_data_validation'] = True
            
            # Generate summary
            analysis['summary'] = self._generate_summary(analysis)
            
            wb.close()
            return analysis
            
        except Exception as e:
            return {
                'error': str(e),
                'file_name': Path(file_path).name
            }
    
    def _analyze_sheet(self, sheet, sheet_name: str) -> Dict[str, Any]:
        """Analyze individual sheet structure"""
        
        analysis = {
            'name': sheet_name,
            'dimensions': f"{sheet.max_row} rows × {sheet.max_column} columns",
            'max_row': sheet.max_row,
            'max_column': sheet.max_column,
            'has_formulas': False,
            'has_data_validation': False,
            'formula_count': 0,
            'data_rows': 0,
            'empty_rows': 0,
            'colored_cells': [],
            'merged_cells': [],
            'sample_data': []
        }
        
        # Check for merged cells
        if hasattr(sheet, 'merged_cells'):
            analysis['merged_cells'] = [str(mc) for mc in sheet.merged_cells.ranges]
        
        # Analyze rows
        for row_idx in range(1, min(sheet.max_row + 1, 101)):  # First 100 rows
            row_has_data = False
            row_data = []
            
            for col_idx in range(1, min(sheet.max_column + 1, 15)):  # First 15 columns
                cell = sheet.cell(row=row_idx, column=col_idx)
                
                # Check for formulas
                if cell.data_type == 'f':
                    analysis['has_formulas'] = True
                    analysis['formula_count'] += 1
                
                # Check for data validation
                if hasattr(cell, 'data_validation') and cell.data_validation:
                    analysis['has_data_validation'] = True
                
                # Check for colored cells (yellow/green for templates)
                if cell.fill and cell.fill.start_color:
                    color = cell.fill.start_color.rgb
                    if color and color != '00000000':
                        if color in ['FFFFFF00', 'FFFF00', 'FFFF0000']:  # Yellow
                            analysis['colored_cells'].append({
                                'cell': f"{cell.column_letter}{cell.row}",
                                'color': 'yellow',
                                'value': str(cell.value)[:30] if cell.value else ''
                            })
                        elif color in ['FF90EE90', '90EE90', 'FF00FF00']:  # Green
                            analysis['colored_cells'].append({
                                'cell': f"{cell.column_letter}{cell.row}",
                                'color': 'green',
                                'value': str(cell.value)[:30] if cell.value else ''
                            })
                
                # Collect sample data
                if cell.value is not None:
                    row_has_data = True
                    value_str = str(cell.value)[:40]
                    row_data.append(value_str)
                else:
                    row_data.append('')
            
            if row_has_data:
                analysis['data_rows'] += 1
                if len(analysis['sample_data']) < 20:  # Keep first 20 data rows
                    analysis['sample_data'].append({
                        'row': row_idx,
                        'data': row_data
                    })
            else:
                analysis['empty_rows'] += 1
        
        return analysis
    
    def _generate_summary(self, analysis: Dict) -> Dict[str, Any]:
        """Generate summary statistics"""
        
        total_data_rows = sum(s['data_rows'] for s in analysis['sheets'].values())
        total_formulas = sum(s['formula_count'] for s in analysis['sheets'].values())
        total_colored = sum(len(s['colored_cells']) for s in analysis['sheets'].values())
        
        # Detect likely sheet types
        sheet_types = {}
        for name, sheet in analysis['sheets'].items():
            if 'meas' in name.lower() or 'measurement' in name.lower():
                sheet_types[name] = 'Measurement Sheet'
            elif 'abs' in name.lower() or 'abstract' in name.lower():
                sheet_types[name] = 'Abstract Sheet'
            elif 'tech' in name.lower() or 'report' in name.lower():
                sheet_types[name] = 'Technical Report'
            elif 'gen' in name.lower() and 'abs' in name.lower():
                sheet_types[name] = 'General Abstract'
            else:
                sheet_types[name] = 'Data Sheet'
        
        return {
            'total_data_rows': total_data_rows,
            'total_formulas': total_formulas,
            'total_colored_cells': total_colored,
            'sheet_types': sheet_types,
            'is_template': total_colored > 0,
            'is_estimate': any('meas' in s.lower() or 'abs' in s.lower() 
                              for s in analysis['sheet_names']),
            'complexity': 'High' if total_formulas > 100 else 'Medium' if total_formulas > 20 else 'Low'
        }
    
    def get_import_recommendations(self, analysis: Dict) -> List[str]:
        """Generate recommendations for importing the file"""
        
        recommendations = []
        
        if analysis.get('error'):
            recommendations.append("⚠️ File has errors - may not import correctly")
            return recommendations
        
        summary = analysis.get('summary', {})
        
        # Check if it's a template
        if summary.get('is_template'):
            recommendations.append("🎨 This appears to be a template (has colored cells)")
            recommendations.append("💡 Use 'Dynamic Template' import mode")
        
        # Check if it's an estimate
        if summary.get('is_estimate'):
            recommendations.append("📊 This appears to be an estimate file")
            recommendations.append("💡 Use 'Standard Estimate' import mode")
        
        # Check complexity
        if summary.get('complexity') == 'High':
            recommendations.append("⚠️ Complex file with many formulas")
            recommendations.append("💡 Import may take longer - formulas will be preserved")
        
        # Check for named ranges
        if analysis.get('has_named_ranges'):
            recommendations.append("✅ File has named ranges - good for templates")
        
        # Check for data validation
        if analysis.get('has_data_validation'):
            recommendations.append("✅ File has data validation rules")
        
        # Sheet-specific recommendations
        sheet_types = summary.get('sheet_types', {})
        if 'Measurement Sheet' in sheet_types.values():
            recommendations.append("📏 Measurement sheets detected")
        if 'Abstract Sheet' in sheet_types.values():
            recommendations.append("💰 Abstract sheets detected")
        
        if not recommendations:
            recommendations.append("✅ File looks good for import")
            recommendations.append("💡 Use 'Auto-detect' import mode")
        
        return recommendations

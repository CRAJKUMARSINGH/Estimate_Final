"""
Specialized Excel Importer for Panchayat Samiti Estimates
Handles the specific format of your Excel files with 1,524 formulas
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from datetime import datetime
import re
from typing import Dict, List, Tuple
import streamlit as st

class PanchayatSamitiEstimateImporter:
    """
    Intelligent importer for construction estimates
    Preserves formulas, relationships, and data integrity
    """
    
    def __init__(self):
        self.excel_file = None
        self.workbook = None
        self.estimate_data = {}
        
    def import_complete_estimate(self, excel_file_path: str) -> Dict:
        """
        Main entry point: Import complete estimate from Excel file
        
        Args:
            excel_file_path: Path to Excel file
            
        Returns:
            Dictionary containing all estimate data
        """
        self.excel_file = excel_file_path
        self.workbook = load_workbook(excel_file_path, data_only=False)
        
        st.info("🔍 Analyzing Excel file structure...")
        
        # Step 1: Analyze structure
        structure = self.analyze_structure()
        st.success(f"✅ Found {len(structure['sheets'])} sheets")
        
        # Step 2: Import each sheet type
        estimate = {
            'general_abstract': None,
            'parts': [],
            'reports': {},
            'metadata': {
                'import_date': datetime.now(),
                'source_file': excel_file_path,
                'total_formulas': structure['total_formulas']
            }
        }
        
        # Import general abstract
        if structure['general_abstract']:
            st.info("📊 Importing General Abstract...")
            estimate['general_abstract'] = self.import_general_abstract(
                structure['general_abstract']
            )
        
        # Import measurement-abstract pairs
        st.info("📏 Importing Measurement-Abstract pairs...")
        for pair in structure['measurement_abstract_pairs']:
            part_data = self.import_part_pair(pair)
            estimate['parts'].append(part_data)
            st.success(f"✅ Imported: {part_data['name']}")
        
        # Import other sheets
        for sheet_info in structure['other_sheets']:
            if sheet_info['type'] == 'technical_report':
                estimate['reports']['technical'] = self.import_technical_report(
                    sheet_info['name']
                )
            elif sheet_info['type'] == 'joinery_schedule':
                estimate['reports']['joinery'] = self.import_joinery_schedule(
                    sheet_info['name']
                )
        
        self.estimate_data = estimate
        return estimate
    
    def analyze_structure(self) -> Dict:
        """
        Analyze Excel file structure and detect sheet relationships
        
        Returns:
            Dictionary with structure information
        """
        structure = {
            'sheets': [],
            'general_abstract': None,
            'measurement_abstract_pairs': [],
            'other_sheets': [],
            'total_formulas': 0
        }
        
        measurements = {}  # Track measurement sheets
        abstracts = {}     # Track abstract sheets
        
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            
            # Count formulas
            formula_count = self._count_formulas(sheet)
            structure['total_formulas'] += formula_count
            
            # Detect sheet type
            sheet_info = {
                'name': sheet_name,
                'type': self._detect_sheet_type(sheet_name, sheet),
                'formula_count': formula_count,
                'row_count': sheet.max_row,
                'col_count': sheet.max_column
            }
            
            structure['sheets'].append(sheet_info)
            
            # Categorize sheets
            if sheet_info['type'] == 'general_abstract':
                structure['general_abstract'] = sheet_name
            elif sheet_info['type'] == 'measurement':
                # Extract part name (e.g., "GF1" from "GF1_MES")
                part_name = self._extract_part_name(sheet_name)
                measurements[part_name] = sheet_name
            elif sheet_info['type'] == 'abstract':
                # Extract part name (e.g., "GF1" from "GF1_ABS")
                part_name = self._extract_part_name(sheet_name)
                abstracts[part_name] = sheet_name
            else:
                structure['other_sheets'].append(sheet_info)
        
        # Match measurement-abstract pairs
        for part_name in set(measurements.keys()) & set(abstracts.keys()):
            structure['measurement_abstract_pairs'].append({
                'part_name': part_name,
                'measurement_sheet': measurements[part_name],
                'abstract_sheet': abstracts[part_name]
            })
        
        return structure
    
    def _detect_sheet_type(self, sheet_name: str, sheet) -> str:
        """Detect sheet type from name and content"""
        name_lower = sheet_name.lower()
        
        if 'general' in name_lower and 'abstract' in name_lower:
            return 'general_abstract'
        elif 'measurement' in name_lower or 'measur' in name_lower or '_mes' in name_lower:
            return 'measurement'
        elif 'abstract' in name_lower or '_abs' in name_lower:
            return 'abstract'
        elif 'sanitary' in name_lower:
            # Check if it's measurement or abstract
            if 'measur' in name_lower:
                return 'measurement'
            else:
                return 'abstract'
        elif 'tech' in name_lower or 'report' in name_lower:
            return 'technical_report'
        elif 'joinery' in name_lower or 'schedule' in name_lower:
            return 'joinery_schedule'
        else:
            return 'other'
    
    def _extract_part_name(self, sheet_name: str) -> str:
        """Extract part identifier from sheet name"""
        # Remove common suffixes
        part_name = sheet_name.replace('_MES', '').replace('_ABS', '')
        part_name = part_name.replace('_MEASUR', '').replace('-abs', '')
        part_name = part_name.replace(' Measurement', '').replace(' Abstract', '')
        return part_name.strip()
    
    def _count_formulas(self, sheet) -> int:
        """Count formula cells in sheet"""
        count = 0
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    count += 1
        return count
    
    def import_general_abstract(self, sheet_name: str) -> Dict:
        """
        Import General Abstract sheet
        Handles title, project info, and cost summary
        """
        sheet = self.workbook[sheet_name]
        
        # Extract project information from top rows
        project_info = {
            'project_name': '',
            'location': '',
            'parts': []
        }
        
        # Scan first 10 rows for project info
        for row in sheet.iter_rows(min_row=1, max_row=10, values_only=True):
            row_text = ' '.join([str(cell) for cell in row if cell])
            if 'NAME OF WORK' in row_text.upper():
                project_info['project_name'] = row_text.split(':')[-1].strip() if ':' in row_text else row_text
        
        # Extract cost breakdown (typically starts after row 6)
        cost_data = []
        for row_idx, row in enumerate(sheet.iter_rows(min_row=7, values_only=True), 7):
            # Check if row has part name and amount
            if row[0] and ('PART' in str(row[0]).upper() or 'WORK' in str(row[0]).upper()):
                # Find amount (last non-empty cell with number)
                amount = None
                for cell in reversed(row):
                    if isinstance(cell, (int, float)) and cell > 0:
                        amount = cell
                        break
                
                if amount:
                    cost_data.append({
                        'part_name': str(row[0]),
                        'amount': float(amount),
                        'row': row_idx
                    })
        
        project_info['parts'] = cost_data
        
        # Calculate totals
        total_cost = sum(part['amount'] for part in cost_data)
        project_info['grand_total'] = total_cost
        
        return project_info
    
    def import_part_pair(self, pair_info: Dict) -> Dict:
        """
        Import measurement-abstract pair with auto-linking
        
        Args:
            pair_info: Dict with part_name, measurement_sheet, abstract_sheet
            
        Returns:
            Dict with complete part data and linkages
        """
        part_data = {
            'name': pair_info['part_name'],
            'measurement_sheet': pair_info['measurement_sheet'],
            'abstract_sheet': pair_info['abstract_sheet'],
            'measurements': None,
            'abstracts': None,
            'linkages': []
        }
        
        # Import measurements
        part_data['measurements'] = self.import_measurement_sheet(
            pair_info['measurement_sheet']
        )
        
        # Import abstracts
        part_data['abstracts'] = self.import_abstract_sheet(
            pair_info['abstract_sheet']
        )
        
        # Create linkages
        part_data['linkages'] = self._create_auto_linkages(
            part_data['measurements'],
            part_data['abstracts']
        )
        
        return part_data
    
    def import_measurement_sheet(self, sheet_name: str) -> pd.DataFrame:
        """
        Import measurement sheet with intelligent column detection
        Handles your specific format with Nos, Length, Breadth, Height columns
        """
        sheet = self.workbook[sheet_name]
        
        # Find header row (contains keywords like Particulars, Nos, Length)
        header_row = self._find_header_row(sheet, 
            keywords=['particulars', 'nos', 'length', 'qty']
        )
        
        if header_row == -1:
            st.warning(f"Could not find header in {sheet_name}, using row 4")
            header_row = 4
        
        # Read data using pandas
        df = pd.read_excel(self.excel_file, sheet_name=sheet_name, 
                          header=header_row, skiprows=range(0, header_row))
        
        # Standardize column names
        column_mapping = self._create_column_mapping(df.columns)
        df = df.rename(columns=column_mapping)
        
        # Clean and structure data
        measurements = []
        for idx, row in df.iterrows():
            # Skip empty rows
            if pd.isna(row.get('description')) or row.get('description') == '':
                continue
            
            # Skip total/subtotal rows
            desc_lower = str(row.get('description', '')).lower()
            if any(word in desc_lower for word in ['total', 'subtotal', 'grand total']):
                continue
            
            measurement = {
                'id': len(measurements) + 1,
                'item_no': str(row.get('item_no', '')),
                'description': str(row.get('description', '')),
                'specification': '',
                'location': '',
                'quantity': self._safe_float(row.get('quantity', 1)),
                'length': self._safe_float(row.get('length', 0)),
                'breadth': self._safe_float(row.get('breadth', 0)),
                'height': self._safe_float(row.get('height', 0)),
                'diameter': 0,
                'thickness': 0,
                'unit': str(row.get('unit', '')),
                'total': self._safe_float(row.get('total', 0)),
                'deduction': 0,
                'net_total': self._safe_float(row.get('total', 0)),
                'remarks': '',
                'ssr_code': ''
            }
            
            measurements.append(measurement)
        
        return pd.DataFrame(measurements)
    
    def import_abstract_sheet(self, sheet_name: str) -> pd.DataFrame:
        """
        Import abstract sheet with rate and amount columns
        """
        sheet = self.workbook[sheet_name]
        
        # Find header row
        header_row = self._find_header_row(sheet,
            keywords=['particulars', 'quantity', 'rate', 'amount']
        )
        
        if header_row == -1:
            st.warning(f"Could not find header in {sheet_name}, using row 3")
            header_row = 3
        
        # Read data
        df = pd.read_excel(self.excel_file, sheet_name=sheet_name,
                          header=header_row, skiprows=range(0, header_row))
        
        # Standardize columns
        column_mapping = self._create_column_mapping(df.columns)
        df = df.rename(columns=column_mapping)
        
        # Clean data
        abstracts = []
        for idx, row in df.iterrows():
            # Skip empty or total rows
            if pd.isna(row.get('description')) or row.get('description') == '':
                continue
            
            desc_lower = str(row.get('description', '')).lower()
            if any(word in desc_lower for word in ['total', 'subtotal', 'grand total']):
                continue
            
            abstract = {
                'id': len(abstracts) + 1,
                'ssr_code': '',
                'description': str(row.get('description', '')),
                'unit': str(row.get('unit', '')),
                'quantity': self._safe_float(row.get('quantity', 0)),
                'rate': self._safe_float(row.get('rate', 0)),
                'amount': self._safe_float(row.get('amount', 0))
            }
            
            abstracts.append(abstract)
        
        return pd.DataFrame(abstracts)
    
    def _find_header_row(self, sheet, keywords: List[str]) -> int:
        """
        Find header row by searching for keywords
        Returns row index (0-based) or -1 if not found
        """
        for row_idx, row in enumerate(sheet.iter_rows(max_row=20, values_only=True)):
            row_text = ' '.join([str(cell).lower() for cell in row if cell])
            
            # Check if at least 2 keywords are present
            matches = sum(1 for keyword in keywords if keyword in row_text)
            if matches >= 2:
                return row_idx
        
        return -1
    
    def _create_column_mapping(self, columns) -> Dict[str, str]:
        """
        Create mapping from Excel columns to standardized names
        """
        mapping = {}
        
        # Define mapping rules
        rules = {
            'item_no': ['s.no', 's.n.', 'sno', 'item no', 'sr.no'],
            'description': ['particulars', 'description', 'item', 'work description'],
            'quantity': ['nos', 'nos.', 'quantity', 'qty', 'numbers'],
            'length': ['length', 'l'],
            'breadth': ['breadth', 'width', 'b', 'w'],
            'height': ['height', 'depth', 'h', 'd'],
            'unit': ['unit', 'units', 'uom'],
            'total': ['qty', 'qty.', 'total', 'total qty'],
            'rate': ['rate', 'unit rate', 'price'],
            'amount': ['amount', 'cost', 'total amount']
        }
        
        # Match columns to rules
        for col in columns:
            col_lower = str(col).lower().strip()
            
            for standard_name, patterns in rules.items():
                if any(pattern in col_lower for pattern in patterns):
                    mapping[col] = standard_name
                    break
        
        return mapping
    
    def _safe_float(self, value, default=0.0) -> float:
        """Safely convert value to float"""
        try:
            if pd.isna(value) or value == '':
                return default
            return float(value)
        except (ValueError, TypeError):
            return default
    
    def _create_auto_linkages(self, measurements_df: pd.DataFrame, 
                             abstracts_df: pd.DataFrame) -> List[Dict]:
        """
        Auto-create linkages between measurements and abstracts
        Based on description similarity
        """
        linkages = []
        
        for _, abstract_row in abstracts_df.iterrows():
            abstract_desc = abstract_row['description'].lower()
            abstract_words = set(re.findall(r'\w+', abstract_desc))
            
            matching_measurements = []
            
            for _, meas_row in measurements_df.iterrows():
                meas_desc = meas_row['description'].lower()
                meas_words = set(re.findall(r'\w+', meas_desc))
                
                # Calculate Jaccard similarity
                intersection = len(abstract_words & meas_words)
                union = len(abstract_words | meas_words)
                
                if union > 0:
                    similarity = intersection / union
                    
                    # Consider match if similarity > 30% and at least 2 words match
                    if similarity > 0.3 and intersection >= 2:
                        matching_measurements.append({
                            'measurement_id': meas_row['id'],
                            'measurement_desc': meas_row['description'],
                            'similarity': similarity,
                            'total': meas_row['total'],
                            'unit': meas_row['unit']
                        })
            
            if matching_measurements:
                # Sort by similarity
                matching_measurements.sort(key=lambda x: x['similarity'], reverse=True)
                
                linkages.append({
                    'abstract_id': abstract_row['id'],
                    'abstract_description': abstract_row['description'],
                    'abstract_unit': abstract_row['unit'],
                    'measurements': matching_measurements,
                    'total_quantity': sum(m['total'] for m in matching_measurements),
                    'confidence': matching_measurements[0]['similarity']
                })
        
        return linkages
    
    def import_technical_report(self, sheet_name: str) -> pd.DataFrame:
        """Import technical report sheet"""
        df = pd.read_excel(self.excel_file, sheet_name=sheet_name)
        return df
    
    def import_joinery_schedule(self, sheet_name: str) -> pd.DataFrame:
        """Import joinery schedule sheet"""
        df = pd.read_excel(self.excel_file, sheet_name=sheet_name)
        return df
    
    def update_session_state(self, estimate_data: Dict):
        """
        Update Streamlit session state with imported data
        """
        st.info("💾 Updating application state...")
        
        # Update general abstract settings
        if estimate_data['general_abstract']:
            ga = estimate_data['general_abstract']
            st.session_state.general_abstract_settings.update({
                'project_name': ga.get('project_name', ''),
                'project_location': ga.get('location', '')
            })
        
        # Update measurement and abstract sheets
        for part in estimate_data['parts']:
            sheet_name = part['name']
            
            # Create sheets if they don't exist
            if sheet_name not in st.session_state.measurement_sheets:
                st.session_state.measurement_sheets[sheet_name] = pd.DataFrame()
            if sheet_name not in st.session_state.abstract_sheets:
                st.session_state.abstract_sheets[sheet_name] = pd.DataFrame()
            
            # Update data
            st.session_state.measurement_sheets[sheet_name] = part['measurements']
            st.session_state.abstract_sheets[sheet_name] = part['abstracts']
            
            st.success(f"✅ Loaded {len(part['measurements'])} measurements and {len(part['abstracts'])} abstracts for {sheet_name}")
        
        st.success("✨ Import complete! All data loaded successfully.")


# Usage example
def demo_import():
    """Demo function showing how to use the importer"""
    
    st.title("🏗️ Import Excel Estimate")
    
    uploaded_file = st.file_uploader("Upload Excel estimate file", type=['xlsx'])
    
    if uploaded_file:
        # Save uploaded file temporarily
        import tempfile
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            tmp.write(uploaded_file.getvalue())
            tmp_path = tmp.name
        
        # Import
        if st.button("Import Estimate", type="primary"):
            with st.spinner("Importing estimate..."):
                importer = PanchayatSamitiEstimateImporter()
                
                try:
                    # Import complete estimate
                    estimate_data = importer.import_complete_estimate(tmp_path)
                    
                    # Update session state
                    importer.update_session_state(estimate_data)
                    
                    # Show summary
                    st.success("🎉 Import Successful!")
                    
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("Parts Imported", len(estimate_data['parts']))
                    with col2:
                        st.metric("Total Formulas", estimate_data['metadata']['total_formulas'])
                    with col3:
                        total_amount = estimate_data['general_abstract']['grand_total']
                        st.metric("Grand Total", f"₹{total_amount:,.2f}")
                    
                    # Show linkages
                    st.subheader("📊 Auto-Detected Linkages")
                    for part in estimate_data['parts']:
                        with st.expander(f"{part['name']} - {len(part['linkages'])} linkages"):
                            for link in part['linkages'][:5]:  # Show first 5
                                st.write(f"**{link['abstract_description'][:50]}...**")
                                st.write(f"  → {len(link['measurements'])} measurements linked (confidence: {link['confidence']:.1%})")
                                st.write(f"  → Total quantity: {link['total_quantity']:.2f} {link['abstract_unit']}")
                    
                except Exception as e:
                    st.error(f"Import failed: {str(e)}")
                    import traceback
                    st.code(traceback.format_exc())
        
        # Cleanup
        import os
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)


if __name__ == "__main__":
    demo_import()

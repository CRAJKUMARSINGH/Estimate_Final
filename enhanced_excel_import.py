#!/usr/bin/env python3
"""
Enhanced Excel Import System for Construction Estimation
Handles importing estimates from attached_assets folder and SSR data
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os
import sys
from pathlib import Path
import json
from datetime import datetime
import streamlit as st
import glob

class EnhancedExcelImporter:
    def __init__(self, base_path="."):
        self.base_path = Path(base_path)
        self.attached_assets_path = self.base_path / "attached_assets"
        self.session_state = st.session_state if 'st' in globals() else {}
        
    def find_estimate_files(self, pattern="att*.xlsx"):
        """Find estimate files in attached_assets folder matching pattern"""
        if not self.attached_assets_path.exists():
            print(f"❌ Attached assets folder not found: {self.attached_assets_path}")
            return []
        
        search_pattern = self.attached_assets_path / pattern
        files = list(glob.glob(str(search_pattern)))
        return sorted(files)
    
    def import_estimate_from_attached(self, file_path=None):
        """Import estimate from attached assets folder"""
        if file_path is None:
            # Find files matching pattern
            files = self.find_estimate_files()
            if not files:
                st.warning("No estimate files found in attached_assets folder matching pattern 'att*.xlsx'")
                return False
            
            # If only one file, use it directly
            if len(files) == 1:
                file_path = files[0]
            else:
                # Let user select file in Streamlit
                if 'st' in globals():
                    selected_file = st.selectbox(
                        "Select estimate file to import:",
                        files,
                        format_func=lambda x: os.path.basename(x)
                    )
                    file_path = selected_file
                else:
                    print("Available files:")
                    for i, file in enumerate(files, 1):
                        print(f"{i}. {os.path.basename(file)}")
                    choice = input("Enter file number: ")
                    try:
                        file_path = files[int(choice) - 1]
                    except (ValueError, IndexError):
                        print("Invalid selection")
                        return False
        
        # Import the selected file
        return self.import_excel_file(file_path)
    
    def import_excel_file(self, file_path):
        """Import Excel file and extract estimate data"""
        try:
            print(f"🔍 Importing estimate from: {os.path.basename(file_path)}")
            
            # Load workbook
            wb = load_workbook(file_path, data_only=True)
            
            # Extract data from sheets
            estimate_data = {
                'general_abstract': None,
                'abstract_sheets': {},
                'measurement_sheets': {},
                'ssr_data': None,
                'metadata': {
                    'file_name': os.path.basename(file_path),
                    'import_date': datetime.now().isoformat(),
                    'sheet_count': len(wb.sheetnames)
                }
            }
            
            # Process each sheet
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet_data = self.extract_sheet_data(ws)
                
                # Categorize sheet
                if "general abstract" in sheet_name.lower():
                    estimate_data['general_abstract'] = {
                        'name': sheet_name,
                        'data': sheet_data
                    }
                elif "abstract of cost" in sheet_name.lower():
                    part_name = sheet_name.replace("Abstract of Cost", "").strip()
                    estimate_data['abstract_sheets'][part_name] = {
                        'name': sheet_name,
                        'data': sheet_data
                    }
                elif "measurement" in sheet_name.lower():
                    part_name = sheet_name.replace("Measurement", "").strip()
                    estimate_data['measurement_sheets'][part_name] = {
                        'name': sheet_name,
                        'data': sheet_data
                    }
                elif "ssr" in sheet_name.lower() or "schedule" in sheet_name.lower():
                    estimate_data['ssr_data'] = {
                        'name': sheet_name,
                        'data': sheet_data
                    }
            
            # Update session state if in Streamlit
            if 'st' in globals():
                st.session_state.estimate_data = estimate_data
                st.success(f"✅ Estimate imported successfully from {os.path.basename(file_path)}!")
                return True
            else:
                print("✅ Estimate imported successfully!")
                return estimate_data
                
        except Exception as e:
            error_msg = f"❌ Error importing estimate: {str(e)}"
            if 'st' in globals():
                st.error(error_msg)
            else:
                print(error_msg)
            return False
    
    def extract_sheet_data(self, worksheet):
        """Extract data from Excel worksheet"""
        data = []
        headers = []
        
        # Get headers from first row
        for col in range(1, worksheet.max_column + 1):
            headers.append(worksheet.cell(row=1, column=col).value)
        
        # Get data rows
        for row in range(1, worksheet.max_row + 1):
            row_data = []
            for col in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=row, column=col).value
                row_data.append(cell_value)
            data.append(row_data)
        
        return {
            'headers': headers,
            'data': data,
            'dimensions': {
                'rows': worksheet.max_row,
                'columns': worksheet.max_column
            }
        }
    
    def add_new_measurement_line(self, sheet_name, item_data):
        """Add new line to measurement sheet"""
        # This would typically interact with Excel directly
        # For Streamlit, we'll update the session state
        if 'st' in globals() and 'measurements' in st.session_state:
            # Add to existing measurements
            new_id = len(st.session_state.measurements) + 1
            new_item = {
                'id': new_id,
                'item_no': item_data.get('item_no', str(new_id)),
                'description': item_data.get('description', ''),
                'quantity': item_data.get('quantity', 0),
                'length': item_data.get('length', 0),
                'breadth': item_data.get('breadth', 0),
                'height': item_data.get('height', 0),
                'unit': item_data.get('unit', 'Cum'),
                'total': item_data.get('quantity', 0) * item_data.get('length', 1) * 
                         item_data.get('breadth', 1) * item_data.get('height', 1),
                'ssr_code': item_data.get('ssr_code', '')
            }
            
            # Convert to DataFrame and append
            import pandas as pd
            new_df = pd.DataFrame([new_item])
            st.session_state.measurements = pd.concat([
                st.session_state.measurements, 
                new_df
            ], ignore_index=True)
            
            st.success("✅ New measurement line added successfully!")
            return True
        return False
    
    def update_all_measurements(self):
        """Update all measurement calculations"""
        if 'st' in globals() and 'measurements' in st.session_state:
            # Recalculate totals
            st.session_state.measurements['total'] = (
                st.session_state.measurements['quantity'] * 
                st.session_state.measurements['length'] * 
                st.session_state.measurements['breadth'] * 
                st.session_state.measurements['height']
            ).round(2)
            
            st.success("✅ All measurements updated successfully!")
            return True
        return False
    
    def import_ssr_from_excel(self, file_path=None):
        """Import SSR data from Excel file"""
        if file_path is None:
            # Find SSR files
            ssr_files = self.find_estimate_files("*ssr*.xlsx")
            if not ssr_files:
                # Try general estimate files
                ssr_files = self.find_estimate_files("*.xlsx")
            
            if not ssr_files:
                st.warning("No SSR files found in attached_assets folder")
                return False
            
            # Let user select file
            if 'st' in globals():
                selected_file = st.selectbox(
                    "Select SSR file to import:",
                    ssr_files,
                    format_func=lambda x: os.path.basename(x),
                    key="ssr_file_select"
                )
                file_path = selected_file
            else:
                print("Available files:")
                for i, file in enumerate(ssr_files, 1):
                    print(f"{i}. {os.path.basename(file)}")
                choice = input("Enter file number: ")
                try:
                    file_path = ssr_files[int(choice) - 1]
                except (ValueError, IndexError):
                    print("Invalid selection")
                    return False
        
        try:
            # Read SSR data
            ssr_df = pd.read_excel(file_path, sheet_name=None)  # Read all sheets
            
            # Find SSR sheet (look for sheet with SSR-like data)
            ssr_sheet = None
            for sheet_name, df in ssr_df.items():
                if 'ssr' in sheet_name.lower() or 'schedule' in sheet_name.lower():
                    ssr_sheet = df
                    break
            
            # If no specific SSR sheet found, use first sheet
            if ssr_sheet is None and ssr_df:
                ssr_sheet = list(ssr_df.values())[0]
            
            if ssr_sheet is not None:
                # Update session state
                if 'st' in globals():
                    st.session_state.ssr_items = ssr_sheet
                    st.success(f"✅ SSR data imported successfully from {os.path.basename(file_path)}!")
                    return True
                else:
                    print("✅ SSR data imported successfully!")
                    return ssr_sheet
            else:
                error_msg = "❌ No valid SSR data found in file"
                if 'st' in globals():
                    st.error(error_msg)
                else:
                    print(error_msg)
                return False
                
        except Exception as e:
            error_msg = f"❌ Error importing SSR data: {str(e)}"
            if 'st' in globals():
                st.error(error_msg)
            else:
                print(error_msg)
            return False
    
    def import_whole_estimate(self, file_path=None):
        """Import entire estimate as a whole"""
        return self.import_estimate_from_attached(file_path)
    
    def display_import_summary(self, estimate_data):
        """Display summary of imported estimate"""
        if 'st' in globals():
            st.subheader("📋 Import Summary")
            
            # File information
            st.write(f"**File:** {estimate_data['metadata']['file_name']}")
            st.write(f"**Import Date:** {estimate_data['metadata']['import_date']}")
            st.write(f"**Total Sheets:** {estimate_data['metadata']['sheet_count']}")
            
            # Sheet breakdown
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("General Abstract", 1 if estimate_data['general_abstract'] else 0)
            with col2:
                st.metric("Abstract Sheets", len(estimate_data['abstract_sheets']))
            with col3:
                st.metric("Measurement Sheets", len(estimate_data['measurement_sheets']))
            with col4:
                st.metric("SSR Data", 1 if estimate_data['ssr_data'] else 0)
            
            # Show sheet names
            if estimate_data['general_abstract']:
                st.write(f"📊 General Abstract: {estimate_data['general_abstract']['name']}")
            
            if estimate_data['abstract_sheets']:
                st.write("💰 Abstract Sheets:")
                for part_name, sheet_info in estimate_data['abstract_sheets'].items():
                    st.write(f"   • {sheet_info['name']}")
            
            if estimate_data['measurement_sheets']:
                st.write("📏 Measurement Sheets:")
                for part_name, sheet_info in estimate_data['measurement_sheets'].items():
                    st.write(f"   • {sheet_info['name']}")
            
            if estimate_data['ssr_data']:
                st.write(f"📚 SSR Data: {estimate_data['ssr_data']['name']}")
        else:
            print("=== Import Summary ===")
            print(f"File: {estimate_data['metadata']['file_name']}")
            print(f"Import Date: {estimate_data['metadata']['import_date']}")
            print(f"Total Sheets: {estimate_data['metadata']['sheet_count']}")
            print(f"General Abstract: {1 if estimate_data['general_abstract'] else 0}")
            print(f"Abstract Sheets: {len(estimate_data['abstract_sheets'])}")
            print(f"Measurement Sheets: {len(estimate_data['measurement_sheets'])}")
            print(f"SSR Data: {1 if estimate_data['ssr_data'] else 0}")

# Streamlit Integration Functions
def streamlit_import_interface():
    """Streamlit interface for enhanced import functionality"""
    st.title("🏗️ Enhanced Excel Import")
    
    # Initialize importer
    importer = EnhancedExcelImporter()
    
    # Import options
    import_option = st.radio(
        "Select Import Type:",
        ["Import from Attached Assets", "Import SSR Data", "Import Whole Estimate"]
    )
    
    if import_option == "Import from Attached Assets":
        if st.button("Import Estimate from att*.xlsx"):
            importer.import_estimate_from_attached()
            
    elif import_option == "Import SSR Data":
        if st.button("Import SSR from Excel"):
            importer.import_ssr_from_excel()
            
    elif import_option == "Import Whole Estimate":
        if st.button("Import Complete Estimate"):
            importer.import_whole_estimate()
    
    # Add new measurement line
    st.subheader("➕ Add New Measurement Line")
    with st.form("add_measurement_line"):
        col1, col2 = st.columns(2)
        with col1:
            item_no = st.text_input("Item No.")
            description = st.text_input("Description")
            unit = st.selectbox("Unit", ["Cum", "Sqm", "Nos", "Kg", "Ton", "Ltr", "LS", "RM"])
        
        with col2:
            quantity = st.number_input("Quantity", value=1.0, min_value=0.0, step=0.01)
            length = st.number_input("Length", value=1.0, min_value=0.0, step=0.01)
            breadth = st.number_input("Breadth", value=1.0, min_value=0.0, step=0.01)
            height = st.number_input("Height", value=1.0, min_value=0.0, step=0.01)
        
        if st.form_submit_button("Add Measurement Line"):
            item_data = {
                'item_no': item_no,
                'description': description,
                'unit': unit,
                'quantity': quantity,
                'length': length,
                'breadth': breadth,
                'height': height
            }
            importer.add_new_measurement_line("Measurements", item_data)
    
    # Update all measurements
    if st.button("🔄 Update All Measurements"):
        importer.update_all_measurements()
    
    # Show import summary if data exists
    if 'estimate_data' in st.session_state:
        importer.display_import_summary(st.session_state.estimate_data)

# Command line interface
def main():
    """Command line interface"""
    importer = EnhancedExcelImporter()
    
    print("🏗️ Enhanced Construction Estimation Excel Importer")
    print("=" * 50)
    
    while True:
        print("\nOptions:")
        print("1. Import estimate from attached_assets (att*.xlsx)")
        print("2. Import SSR data from Excel")
        print("3. Import whole estimate")
        print("4. Add new measurement line")
        print("5. Update all measurements")
        print("6. Exit")
        
        choice = input("\nEnter your choice (1-6): ")
        
        if choice == "1":
            importer.import_estimate_from_attached()
        elif choice == "2":
            importer.import_ssr_from_excel()
        elif choice == "3":
            importer.import_whole_estimate()
        elif choice == "4":
            print("Adding new measurement line...")
            item_data = {
                'item_no': input("Item No.: "),
                'description': input("Description: "),
                'unit': input("Unit (default Cum): ") or "Cum",
                'quantity': float(input("Quantity (default 1): ") or "1"),
                'length': float(input("Length (default 1): ") or "1"),
                'breadth': float(input("Breadth (default 1): ") or "1"),
                'height': float(input("Height (default 1): ") or "1")
            }
            importer.add_new_measurement_line("Measurements", item_data)
        elif choice == "5":
            importer.update_all_measurements()
        elif choice == "6":
            print("Goodbye!")
            break
        else:
            print("Invalid choice. Please try again.")

if __name__ == "__main__":
    # Check if running in Streamlit
    if 'streamlit' in sys.modules or ('st' in globals() and hasattr(st, 'session_state')):
        streamlit_import_interface()
    else:
        main()
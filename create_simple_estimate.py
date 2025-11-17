"""
Create Simple One-Line Measurement Estimate
============================================
Quick estimate creator for single item construction work
"""
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def create_simple_estimate(
    work_name: str,
    item_description: str,
    nos: int,
    length: float,
    breadth: float,
    height: float,
    unit: str,
    rate: float,
    location: str = "Udaipur",
    client: str = "PWD Rajasthan",
    engineer: str = "Er. Rajkumar"
):
    """
    Create a simple one-line measurement estimate
    
    Args:
        work_name: Name of construction work
        item_description: Description of work item
        nos: Number of items
        length: Length in meters
        breadth: Breadth in meters
        height: Height in meters
        unit: Unit of measurement (Cum, Sqm, RM, etc.)
        rate: Rate per unit in INR
        location: Project location
        client: Client name
        engineer: Engineer name
    """
    
    # Calculate quantity
    if height > 0:
        quantity = nos * length * breadth * height
    elif breadth > 0:
        quantity = nos * length * breadth
    elif length > 0:
        quantity = nos * length
    else:
        quantity = nos
    
    # Calculate amount
    amount = quantity * rate
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # ============ SHEET 1: TECHNICAL REPORT ============
    ws_tech = wb.active
    ws_tech.title = "Technical Report"
    
    # Styles
    header_fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Header
    ws_tech['A1'] = f"ESTIMATE FOR {work_name.upper()}"
    ws_tech['A1'].font = title_font
    ws_tech['A1'].alignment = Alignment(horizontal='center')
    ws_tech.merge_cells('A1:F1')
    
    # Project details
    row = 3
    details = [
        ("Name of Work", work_name),
        ("Location", location),
        ("Client", client),
        ("Engineer", engineer),
        ("Date", datetime.now().strftime('%d-%m-%Y'))
    ]
    
    for label, value in details:
        ws_tech[f'A{row}'] = label
        ws_tech[f'A{row}'].font = Font(bold=True)
        ws_tech[f'B{row}'] = ":"
        ws_tech[f'C{row}'] = value
        ws_tech.merge_cells(f'C{row}:F{row}')
        row += 1
    
    # ============ SHEET 2: MEASUREMENTS ============
    ws_meas = wb.create_sheet("Measurements")
    
    # Header
    ws_meas['A1'] = "DETAILS OF MEASUREMENTS"
    ws_meas['A1'].font = title_font
    ws_meas['A1'].alignment = Alignment(horizontal='center')
    ws_meas.merge_cells('A1:H1')
    
    # Column headers
    headers = ['S.N.', 'Particulars', 'Nos', 'Length', 'Breadth', 'Height', 'Qty', 'Unit']
    for col, header in enumerate(headers, 1):
        cell = ws_meas.cell(3, col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Data row
    ws_meas['A4'] = 1
    ws_meas['B4'] = item_description
    ws_meas['C4'] = nos
    ws_meas['D4'] = length
    ws_meas['E4'] = breadth
    ws_meas['F4'] = height
    ws_meas['G4'] = round(quantity, 3)
    ws_meas['H4'] = unit
    
    # Format data row
    for col in range(1, 9):
        cell = ws_meas.cell(4, col)
        cell.border = border
        if col >= 3:  # Numeric columns
            cell.alignment = Alignment(horizontal='right')
    
    # Column widths
    ws_meas.column_dimensions['A'].width = 6
    ws_meas.column_dimensions['B'].width = 40
    ws_meas.column_dimensions['C'].width = 8
    ws_meas.column_dimensions['D'].width = 10
    ws_meas.column_dimensions['E'].width = 10
    ws_meas.column_dimensions['F'].width = 10
    ws_meas.column_dimensions['G'].width = 12
    ws_meas.column_dimensions['H'].width = 8
    
    # ============ SHEET 3: ABSTRACT ============
    ws_abs = wb.create_sheet("Abstract")
    
    # Header
    ws_abs['A1'] = "ABSTRACT OF COST"
    ws_abs['A1'].font = title_font
    ws_abs['A1'].alignment = Alignment(horizontal='center')
    ws_abs.merge_cells('A1:F1')
    
    # Column headers
    headers = ['S.N.', 'Description of Item', 'Quantity', 'Unit', 'Rate (₹)', 'Amount (₹)']
    for col, header in enumerate(headers, 1):
        cell = ws_abs.cell(3, col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Data row
    ws_abs['A4'] = 1
    ws_abs['B4'] = item_description
    ws_abs['C4'] = round(quantity, 3)
    ws_abs['D4'] = unit
    ws_abs['E4'] = rate
    ws_abs['F4'] = round(amount, 2)
    
    # Format data row
    for col in range(1, 7):
        cell = ws_abs.cell(4, col)
        cell.border = border
        if col >= 3:  # Numeric columns
            cell.alignment = Alignment(horizontal='right')
    
    # Total row
    ws_abs['A5'] = ""
    ws_abs['B5'] = "TOTAL"
    ws_abs['B5'].font = Font(bold=True)
    ws_abs['C5'] = ""
    ws_abs['D5'] = ""
    ws_abs['E5'] = ""
    ws_abs['F5'] = round(amount, 2)
    ws_abs['F5'].font = Font(bold=True)
    
    for col in range(1, 7):
        cell = ws_abs.cell(5, col)
        cell.border = border
        cell.fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
    
    # Grand Total
    ws_abs['A7'] = "GRAND TOTAL (in words):"
    ws_abs['A7'].font = Font(bold=True)
    ws_abs.merge_cells('A7:B7')
    
    # Convert amount to words (simplified)
    amount_words = f"Rupees {int(amount):,} only"
    ws_abs['C7'] = amount_words
    ws_abs.merge_cells('C7:F7')
    
    # Column widths
    ws_abs.column_dimensions['A'].width = 6
    ws_abs.column_dimensions['B'].width = 40
    ws_abs.column_dimensions['C'].width = 12
    ws_abs.column_dimensions['D'].width = 8
    ws_abs.column_dimensions['E'].width = 12
    ws_abs.column_dimensions['F'].width = 15
    
    # Save file
    output_dir = Path("generated_estimates")
    output_dir.mkdir(exist_ok=True)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"Simple_Estimate_{timestamp}.xlsx"
    output_path = output_dir / filename
    
    wb.save(output_path)
    
    return {
        'file_path': str(output_path),
        'file_name': filename,
        'quantity': quantity,
        'amount': amount,
        'work_name': work_name
    }


# ============ QUICK EXAMPLES ============

def example_excavation():
    """Example: Excavation work"""
    return create_simple_estimate(
        work_name="Construction of Residential Building",
        item_description="Earth work in excavation in foundation trenches in all types of soil",
        nos=1,
        length=25.0,
        breadth=20.0,
        height=1.5,
        unit="Cum",
        rate=92.00,
        location="Udaipur",
        client="PWD Rajasthan",
        engineer="Er. Rajkumar"
    )


def example_concrete():
    """Example: Concrete work"""
    return create_simple_estimate(
        work_name="Construction of Bridge",
        item_description="Providing and laying M-30 grade concrete in foundation",
        nos=1,
        length=15.0,
        breadth=7.5,
        height=0.25,
        unit="Cum",
        rate=5850.00,
        location="Nathdwara",
        client="PWD Rajasthan",
        engineer="Er. Rajkumar"
    )


def example_plastering():
    """Example: Plastering work"""
    return create_simple_estimate(
        work_name="Construction of School Building",
        item_description="12mm thick cement plaster (1:4) on walls",
        nos=2,
        length=30.0,
        breadth=3.5,
        height=0.0,  # Area calculation
        unit="Sqm",
        rate=185.00,
        location="Udaipur",
        client="Education Department",
        engineer="Er. Rajkumar"
    )


def example_brickwork():
    """Example: Brick work"""
    return create_simple_estimate(
        work_name="Construction of Boundary Wall",
        item_description="Brick work with F.P.S. bricks in cement mortar (1:6)",
        nos=1,
        length=50.0,
        breadth=0.23,
        height=2.5,
        unit="Cum",
        rate=4850.00,
        location="Udaipur",
        client="PWD Rajasthan",
        engineer="Er. Rajkumar"
    )


# ============ MAIN EXECUTION ============

if __name__ == "__main__":
    print("="*80)
    print("SIMPLE ONE-LINE ESTIMATE CREATOR")
    print("="*80)
    print()
    
    # Create examples
    examples = [
        ("Excavation Work", example_excavation),
        ("Concrete Work", example_concrete),
        ("Plastering Work", example_plastering),
        ("Brick Work", example_brickwork)
    ]
    
    print("Creating example estimates...\n")
    
    for name, func in examples:
        result = func()
        print(f"✅ {name}")
        print(f"   File: {result['file_name']}")
        print(f"   Quantity: {result['quantity']:.3f}")
        print(f"   Amount: ₹{result['amount']:,.2f}")
        print()
    
    print("="*80)
    print("✅ All estimates created in: generated_estimates/")
    print("="*80)

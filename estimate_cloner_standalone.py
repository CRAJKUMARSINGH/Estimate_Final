"""
Standalone Estimate Cloner (No Streamlit Dependencies)
"""

from datetime import datetime
from pathlib import Path
from typing import Dict, Optional

import openpyxl
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# PDF generation imports
try:
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm, inch
    from reportlab.platypus import (PageBreak, Paragraph, SimpleDocTemplate,
                                    Spacer, Table, TableStyle)
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    print("Warning: ReportLab not available. PDF export will not work.")


class EstimateClonerStandalone:
    """Clone and modify existing estimates - Standalone version"""
    
    def __init__(self):
        self.source_estimate = None
        self.modifications = []
    
    def load_estimate(self, file_path: str) -> Dict:
        """Load an archived estimate"""
        try:
            wb = load_workbook(file_path, data_only=False)
            
            estimate_data = {
                'file_path': file_path,
                'file_name': Path(file_path).name,
                'sheets': {},
                'workbook': wb
            }
            
            # Load each sheet
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                
                # Convert to DataFrame
                data = []
                for row in sheet.iter_rows(values_only=True):
                    data.append(row)
                
                if data:
                    df = pd.DataFrame(data[1:], columns=data[0])
                    estimate_data['sheets'][sheet_name] = df
            
            self.source_estimate = estimate_data
            return estimate_data
            
        except Exception as e:
            print(f"Error loading estimate: {e}")
            return None
    
    def modify_measurement(self, sheet_name: str, row_index: int, 
                          column: str, new_value: any) -> bool:
        """Modify a measurement value"""
        try:
            if sheet_name in self.source_estimate['sheets']:
                df = self.source_estimate['sheets'][sheet_name].copy()
                df.at[row_index, column] = new_value
                
                # Recalculate if it's a quantity/rate change
                if column in ['Quantity', 'Rate']:
                    if 'Quantity' in df.columns and 'Rate' in df.columns and 'Amount' in df.columns:
                        df['Amount'] = pd.to_numeric(df['Quantity'], errors='coerce') * \
                                      pd.to_numeric(df['Rate'], errors='coerce')
                
                self.source_estimate['sheets'][sheet_name] = df
                
                self.modifications.append({
                    'type': 'modify',
                    'sheet': sheet_name,
                    'row': row_index,
                    'column': column,
                    'value': new_value,
                    'timestamp': datetime.now().isoformat()
                })
                
                return True
        except Exception as e:
            print(f"Error modifying measurement: {e}")
            return False
    
    def add_bsr_item(self, sheet_name: str, item_data: Dict, 
                     position: Optional[int] = None) -> bool:
        """Add a new BSR item to the estimate"""
        try:
            if sheet_name in self.source_estimate['sheets']:
                df = self.source_estimate['sheets'][sheet_name].copy()
                
                # Create new row
                new_row = pd.DataFrame([item_data])
                
                # Insert at position or append
                if position is not None and position < len(df):
                    df = pd.concat([
                        df.iloc[:position],
                        new_row,
                        df.iloc[position:]
                    ]).reset_index(drop=True)
                else:
                    df = pd.concat([df, new_row], ignore_index=True)
                
                # Recalculate amounts
                if 'Quantity' in df.columns and 'Rate' in df.columns and 'Amount' in df.columns:
                    df['Amount'] = pd.to_numeric(df['Quantity'], errors='coerce') * \
                                  pd.to_numeric(df['Rate'], errors='coerce')
                
                self.source_estimate['sheets'][sheet_name] = df
                
                self.modifications.append({
                    'type': 'add',
                    'sheet': sheet_name,
                    'item': item_data,
                    'position': position,
                    'timestamp': datetime.now().isoformat()
                })
                
                return True
        except Exception as e:
            print(f"Error adding BSR item: {e}")
            return False
    
    def recalculate_totals(self, sheet_name: str) -> bool:
        """Recalculate all totals in a sheet"""
        try:
            if sheet_name in self.source_estimate['sheets']:
                df = self.source_estimate['sheets'][sheet_name].copy()
                
                # Recalculate amounts
                if 'Quantity' in df.columns and 'Rate' in df.columns:
                    if 'Amount' not in df.columns:
                        df['Amount'] = 0
                    
                    df['Amount'] = pd.to_numeric(df['Quantity'], errors='coerce').fillna(0) * \
                                  pd.to_numeric(df['Rate'], errors='coerce').fillna(0)
                
                self.source_estimate['sheets'][sheet_name] = df
                return True
        except Exception as e:
            print(f"Error recalculating totals: {e}")
            return False
    
    def save_as_new_estimate(self, output_path: str, project_info: Dict) -> bool:
        """Save modified estimate as new file"""
        try:
            # Create new workbook
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            # Add each sheet
            for sheet_name, df in self.source_estimate['sheets'].items():
                ws = wb.create_sheet(sheet_name)
                
                # Write data
                for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                    for c_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=r_idx, column=c_idx, value=value)
                        
                        # Style header row
                        if r_idx == 1:
                            cell.font = Font(bold=True, color="FFFFFF")
                            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # Add metadata sheet
            meta_ws = wb.create_sheet("Metadata", 0)
            meta_ws['A1'] = "Project Information"
            meta_ws['A1'].font = Font(bold=True, size=14)
            
            row = 3
            for key, value in project_info.items():
                meta_ws[f'A{row}'] = key
                meta_ws[f'B{row}'] = value
                meta_ws[f'A{row}'].font = Font(bold=True)
                row += 1
            
            # Add modifications log
            row += 2
            meta_ws[f'A{row}'] = "Modifications Log"
            meta_ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            
            for mod in self.modifications:
                meta_ws[f'A{row}'] = f"{mod['type'].upper()}: {mod.get('sheet', '')} - {mod.get('timestamp', '')}"
                row += 1
            
            # Save
            wb.save(output_path)
            
            return True
            
        except Exception as e:
            print(f"Error saving estimate: {e}")
            return False
    
    def export_to_pdf(self, output_path: str, project_info: Dict, use_timestamped_folder: bool = True) -> bool:
        """
        Export estimate to professional A4 PDF
        
        Args:
            output_path: Base path for PDF file
            project_info: Project information dictionary
            use_timestamped_folder: If True, saves in date/time stamped subfolder
        """
        if not REPORTLAB_AVAILABLE:
            print("ReportLab not installed. Run: pip install reportlab")
            return False
        
        try:
            # Create timestamped folder if requested
            if use_timestamped_folder:
                from pathlib import Path

                # Create PDF_Exports folder
                base_dir = Path("PDF_Exports")
                base_dir.mkdir(exist_ok=True)
                
                # Create timestamped subfolder
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                date_folder = datetime.now().strftime('%Y-%m-%d')
                timestamped_folder = base_dir / date_folder / timestamp
                timestamped_folder.mkdir(parents=True, exist_ok=True)
                
                # Update output path
                output_filename = Path(output_path).name
                output_path = str(timestamped_folder / output_filename)
                
                print(f"PDF will be saved to: {output_path}")
            # Create PDF document
            doc = SimpleDocTemplate(
                output_path,
                pagesize=A4,
                rightMargin=1.5*cm,
                leftMargin=1.5*cm,
                topMargin=2*cm,
                bottomMargin=2*cm
            )
            
            # Container for PDF elements
            elements = []
            
            # Styles
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                textColor=colors.HexColor('#1f4e79'),
                spaceAfter=30,
                alignment=TA_CENTER,
                fontName='Helvetica-Bold'
            )
            
            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontSize=12,
                textColor=colors.HexColor('#2d5aa0'),
                spaceAfter=12,
                spaceBefore=12,
                fontName='Helvetica-Bold'
            )
            
            # Title
            elements.append(Paragraph("CONSTRUCTION ESTIMATE", title_style))
            elements.append(Spacer(1, 0.3*cm))
            
            # Project Information Box
            project_data = [
                ['Project Name:', project_info.get('Project Name', 'N/A')],
                ['Location:', project_info.get('Location', 'N/A')],
                ['Client:', project_info.get('Client Name', 'N/A')],
                ['Engineer:', project_info.get('Engineer Name', 'N/A')],
                ['Date:', project_info.get('Date Prepared', 'N/A')],
                ['Estimated Cost:', f"₹ {project_info.get('Estimated Cost', 0):,.2f}"]
            ]
            
            project_table = Table(project_data, colWidths=[4*cm, 12*cm])
            project_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e7e6e6')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
                ('ALIGN', (1, 0), (1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 8),
                ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            
            elements.append(project_table)
            elements.append(Spacer(1, 0.5*cm))
            
            # Add each sheet
            for sheet_name, df in self.source_estimate['sheets'].items():
                # Skip metadata sheet
                if sheet_name.lower() == 'metadata':
                    continue
                
                # Sheet heading
                elements.append(PageBreak())
                elements.append(Paragraph(f"{sheet_name}", heading_style))
                elements.append(Spacer(1, 0.3*cm))
                
                # Prepare data - select key columns
                key_columns = []
                for col in ['Sr No', 'Description', 'Quantity', 'Unit', 'Rate', 'Amount']:
                    if col in df.columns:
                        key_columns.append(col)
                
                # If no standard columns, use first 6
                if not key_columns:
                    key_columns = df.columns.tolist()[:6]
                
                df_display = df[key_columns].copy()
                
                # Convert to list
                table_data = [key_columns]
                
                # Add rows (limit for PDF size)
                max_rows = 40
                for idx, row in df_display.head(max_rows).iterrows():
                    row_data = []
                    for val in row:
                        # Format numbers
                        if isinstance(val, (int, float)):
                            if val > 1000:
                                row_data.append(f'{val:,.2f}')
                            else:
                                row_data.append(f'{val:.2f}')
                        else:
                            row_data.append(str(val)[:40] if val else '')
                    table_data.append(row_data)
                
                if len(df) > max_rows:
                    table_data.append(['...' for _ in range(len(key_columns))])
                
                # Calculate column widths
                available_width = 17*cm
                col_widths = [available_width / len(key_columns)] * len(key_columns)
                
                # Adjust for description column
                if 'Description' in key_columns:
                    desc_idx = key_columns.index('Description')
                    col_widths[desc_idx] = 6*cm
                    remaining = (available_width - 6*cm) / (len(key_columns) - 1)
                    for i in range(len(col_widths)):
                        if i != desc_idx:
                            col_widths[i] = remaining
                
                # Create table
                data_table = Table(table_data, colWidths=col_widths, repeatRows=1)
                data_table.setStyle(TableStyle([
                    # Header
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 8),
                    
                    # Data
                    ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                    ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # Sr No
                    ('ALIGN', (1, 1), (1, -1), 'LEFT'),    # Description
                    ('ALIGN', (2, 1), (-1, -1), 'RIGHT'),  # Numbers
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 7),
                    ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('LEFTPADDING', (0, 0), (-1, -1), 3),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                    ('TOPPADDING', (0, 0), (-1, -1), 2),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
                ]))
                
                elements.append(data_table)
                elements.append(Spacer(1, 0.4*cm))
                
                # Add sheet total
                if 'Amount' in df.columns:
                    try:
                        total = pd.to_numeric(df['Amount'], errors='coerce').sum()
                        summary_data = [[f'{sheet_name} Total:', f'₹ {total:,.2f}']]
                        summary_table = Table(summary_data, colWidths=[10*cm, 6*cm])
                        summary_table.setStyle(TableStyle([
                            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#d9e1f2')),
                            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
                            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                            ('FONTSIZE', (0, 0), (-1, -1), 10),
                            ('BOX', (0, 0), (-1, -1), 1, colors.black),
                            ('LEFTPADDING', (0, 0), (-1, -1), 8),
                            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                            ('TOPPADDING', (0, 0), (-1, -1), 6),
                            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                        ]))
                        elements.append(summary_table)
                    except:
                        pass
            
            # Build PDF
            doc.build(elements)
            
            return True
            
        except Exception as e:
            print(f"Error generating PDF: {e}")
            import traceback
            traceback.print_exc()
            return False

#!/usr/bin/env python3
"""
Excel Estimate File Reader and Display System
Demonstrates the import and structure analysis of construction estimate files
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os
import sys
from pathlib import Path
import json
from datetime import datetime

class EstimateFileReader:
    def __init__(self):
        self.workbook = None
        self.file_path = None
        self.sheet_structure = {}
        self.abstract_sheets = []
        self.measurement_sheets = []
        self.general_abstract = None
        
    def load_estimate_file(self, file_path):
        """Load and analyze Excel estimate file"""
        try:
            print(f"🔍 Loading estimate file: {file_path}")
            self.file_path = file_path
            self.workbook = load_workbook(file_path, data_only=False)
            
            print(f"✅ File loaded successfully!")
            print(f"📊 Total sheets found: {len(self.workbook.sheetnames)}")
            
            # Analyze sheet structure
            self.analyze_sheet_structure()
            return True
            
        except Exception as e:
            print(f"❌ Error loading file: {str(e)}")
            return False
    
    def analyze_sheet_structure(self):
        """Analyze and categorize sheets"""
        print("\n🔍 Analyzing sheet structure...")
        
        for sheet_name in self.workbook.sheetnames:
            sheet_type = self.identify_sheet_type(sheet_name)
            self.sheet_structure[sheet_name] = sheet_type
            
            if sheet_type == "General Abstract":
                self.general_abstract = sheet_name
            elif sheet_type == "Abstract of Cost":
                self.abstract_sheets.append(sheet_name)
            elif sheet_type == "Measurement":
                self.measurement_sheets.append(sheet_name)
        
        print(f"📋 General Abstract: {self.general_abstract}")
        print(f"💰 Abstract sheets: {len(self.abstract_sheets)}")
        print(f"📏 Measurement sheets: {len(self.measurement_sheets)}")
    
    def identify_sheet_type(self, sheet_name):
        """Identify the type of sheet based on name patterns"""
        name_lower = sheet_name.lower()
        
        if "general abstract" in name_lower or "general" in name_lower:
            return "General Abstract"
        elif "abstract of cost" in name_lower or "abstract" in name_lower:
            return "Abstract of Cost"
        elif "measurement" in name_lower:
            return "Measurement"
        elif "ssr" in name_lower or "schedule" in name_lower:
            return "SSR Database"
        else:
            return "Other"
    
    def display_file_overview(self):
        """Display comprehensive file overview"""
        if not self.workbook:
            print("❌ No file loaded!")
            return
        
        print("\n" + "="*80)
        print("📊 CONSTRUCTION ESTIMATE FILE ANALYSIS")
        print("="*80)
        
        print(f"📁 File: {os.path.basename(self.file_path)}")
        print(f"📅 Analysis Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"📋 Total Sheets: {len(self.workbook.sheetnames)}")
        
        # Display sheet categorization
        print("\n📂 SHEET CATEGORIZATION:")
        print("-" * 40)
        
        for sheet_name, sheet_type in self.sheet_structure.items():
            icon = self.get_sheet_icon(sheet_type)
            print(f"{icon} {sheet_name} ({sheet_type})")
        
        # Display linkage analysis
        print("\n🔗 SHEET LINKAGE ANALYSIS:")
        print("-" * 40)
        self.analyze_sheet_linkages()
    
    def get_sheet_icon(self, sheet_type):
        """Get appropriate icon for sheet type"""
        icons = {
            "General Abstract": "📊",
            "Abstract of Cost": "💰",
            "Measurement": "📏",
            "SSR Database": "📚",
            "Other": "📄"
        }
        return icons.get(sheet_type, "📄")
    
    def analyze_sheet_linkages(self):
        """Analyze potential linkages between sheets"""
        # Find paired sheets
        pairs = self.find_sheet_pairs()
        
        if pairs:
            print("🔗 Detected Sheet Pairs:")
            for abstract, measurement in pairs:
                print(f"   💰 {abstract} ↔ 📏 {measurement}")
        else:
            print("⚠️  No clear sheet pairs detected")
        
        # Check for formulas referencing other sheets
        print("\n🧮 Cross-Sheet Formula References:")
        self.analyze_cross_sheet_formulas()
    
    def find_sheet_pairs(self):
        """Find Abstract-Measurement sheet pairs"""
        pairs = []
        
        for abstract_sheet in self.abstract_sheets:
            # Extract part name from abstract sheet
            part_name = self.extract_part_name(abstract_sheet, "abstract")
            
            # Look for corresponding measurement sheet
            for measurement_sheet in self.measurement_sheets:
                measurement_part = self.extract_part_name(measurement_sheet, "measurement")
                
                if part_name and measurement_part and part_name.lower() == measurement_part.lower():
                    pairs.append((abstract_sheet, measurement_sheet))
                    break
        
        return pairs
    
    def extract_part_name(self, sheet_name, sheet_type):
        """Extract part name from sheet name"""
        name_lower = sheet_name.lower()
        
        if sheet_type == "abstract":
            if "abstract of cost" in name_lower:
                return sheet_name.replace("Abstract of Cost", "").strip()
            elif "abstract" in name_lower:
                return sheet_name.replace("Abstract", "").strip()
        elif sheet_type == "measurement":
            if "measurement" in name_lower:
                return sheet_name.replace("Measurement", "").strip()
        
        return None
    
    def analyze_cross_sheet_formulas(self):
        """Analyze formulas that reference other sheets"""
        formula_count = 0
        
        for sheet_name in self.workbook.sheetnames:
            ws = self.workbook[sheet_name]
            sheet_formulas = 0
            
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        # Check if formula references another sheet
                        if "'" in cell.value or "!" in cell.value:
                            sheet_formulas += 1
                            formula_count += 1
            
            if sheet_formulas > 0:
                print(f"   📋 {sheet_name}: {sheet_formulas} cross-sheet formulas")
        
        if formula_count == 0:
            print("   ℹ️  No cross-sheet formulas detected")
    
    def display_sheet_details(self, sheet_name):
        """Display detailed information about a specific sheet"""
        if sheet_name not in self.workbook.sheetnames:
            print(f"❌ Sheet '{sheet_name}' not found!")
            return
        
        ws = self.workbook[sheet_name]
        sheet_type = self.sheet_structure.get(sheet_name, "Unknown")
        
        print(f"\n📋 SHEET DETAILS: {sheet_name}")
        print("=" * 60)
        print(f"🏷️  Type: {sheet_type}")
        print(f"📐 Dimensions: {ws.max_row} rows × {ws.max_column} columns")
        
        # Display data preview
        print(f"\n📊 DATA PREVIEW (First 10 rows):")
        print("-" * 60)
        
        # Read data using pandas for better formatting
        try:
            df = pd.read_excel(self.file_path, sheet_name=sheet_name, nrows=10)
            if not df.empty:
                print(df.to_string(index=False, max_cols=8))
            else:
                print("   (No data found)")
        except Exception as e:
            print(f"   ⚠️  Could not read data: {str(e)}")
        
        # Analyze formulas in this sheet
        self.analyze_sheet_formulas(ws, sheet_name)
    
    def analyze_sheet_formulas(self, ws, sheet_name):
        """Analyze formulas in a specific sheet"""
        formulas = []
        
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formulas.append({
                        'cell': cell.coordinate,
                        'formula': cell.value
                    })
        
        if formulas:
            print(f"\n🧮 FORMULAS FOUND ({len(formulas)} total):")
            print("-" * 40)
            
            # Show first 5 formulas as examples
            for i, formula_info in enumerate(formulas[:5]):
                print(f"   {formula_info['cell']}: {formula_info['formula']}")
            
            if len(formulas) > 5:
                print(f"   ... and {len(formulas) - 5} more formulas")
        else:
            print(f"\n🧮 FORMULAS: None found")
    
    def display_summary_report(self):
        """Display comprehensive summary report"""
        if not self.workbook:
            print("❌ No file loaded!")
            return
        
        print("\n" + "="*80)
        print("📊 ESTIMATE FILE SUMMARY REPORT")
        print("="*80)
        
        # File information
        file_size = os.path.getsize(self.file_path) / 1024  # KB
        print(f"📁 File: {os.path.basename(self.file_path)}")
        print(f"💾 Size: {file_size:.1f} KB")
        print(f"📅 Analysis: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        
        # Structure summary
        print(f"\n📊 STRUCTURE SUMMARY:")
        print(f"   📋 Total Sheets: {len(self.workbook.sheetnames)}")
        print(f"   📊 General Abstract: {'✅' if self.general_abstract else '❌'}")
        print(f"   💰 Abstract Sheets: {len(self.abstract_sheets)}")
        print(f"   📏 Measurement Sheets: {len(self.measurement_sheets)}")
        
        # Linkage analysis
        pairs = self.find_sheet_pairs()
        print(f"   🔗 Paired Sheets: {len(pairs)}")
        
        # Data summary
        total_rows = 0
        total_formulas = 0
        
        for sheet_name in self.workbook.sheetnames:
            ws = self.workbook[sheet_name]
            total_rows += ws.max_row
            
            # Count formulas
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        total_formulas += 1
        
        print(f"\n📊 DATA SUMMARY:")
        print(f"   📝 Total Data Rows: {total_rows}")
        print(f"   🧮 Total Formulas: {total_formulas}")
        
        # Recommendations
        print(f"\n💡 IMPORT RECOMMENDATIONS:")
        if self.general_abstract:
            print("   ✅ General Abstract detected - good structure")
        else:
            print("   ⚠️  No General Abstract found - may need manual setup")
        
        if len(pairs) > 0:
            print(f"   ✅ {len(pairs)} sheet pairs detected - linkages can be automated")
        else:
            print("   ⚠️  No clear sheet pairs - manual linking may be required")
        
        if total_formulas > 0:
            print(f"   ✅ {total_formulas} formulas found - calculations preserved")
        else:
            print("   ⚠️  No formulas detected - may need formula setup")

def main():
    """Main demonstration function"""
    print("🏗️ CONSTRUCTION ESTIMATE FILE READER")
    print("=" * 50)
    
    # Initialize reader
    reader = EstimateFileReader()
    
    # Look for Excel files in attached_assets
    assets_path = Path("attached_assets")
    excel_files = []
    
    if assets_path.exists():
        excel_files = list(assets_path.glob("*.xlsx")) + list(assets_path.glob("*.xls"))
    
    if not excel_files:
        print("❌ No Excel files found in attached_assets folder!")
        return
    
    print(f"📁 Found {len(excel_files)} Excel file(s):")
    for i, file_path in enumerate(excel_files, 1):
        print(f"   {i}. {file_path.name}")
    
    # Load the first Excel file (or XXXX.xlsx if available)
    target_file = None
    for file_path in excel_files:
        if "XXXX" in file_path.name:
            target_file = file_path
            break
    
    if not target_file:
        target_file = excel_files[0]
    
    print(f"\n🔍 Analyzing: {target_file.name}")
    print("-" * 50)
    
    # Load and analyze the file
    if reader.load_estimate_file(str(target_file)):
        # Display comprehensive analysis
        reader.display_file_overview()
        
        # Display details for each sheet
        print(f"\n📋 DETAILED SHEET ANALYSIS:")
        print("=" * 60)
        
        for sheet_name in reader.workbook.sheetnames:
            reader.display_sheet_details(sheet_name)
            print()  # Add spacing between sheets
        
        # Display summary report
        reader.display_summary_report()
        
        # Simulate import process
        print(f"\n🔄 SIMULATING IMPORT PROCESS:")
        print("=" * 50)
        simulate_import_process(reader)
    
    else:
        print("❌ Failed to load estimate file!")

def simulate_import_process(reader):
    """Simulate the VBA import process"""
    print("1️⃣ Clearing existing sheets...")
    print("   ✅ Existing sheets cleared (except General Abstract)")
    
    print("\n2️⃣ Importing General Abstract...")
    if reader.general_abstract:
        print(f"   ✅ Imported: {reader.general_abstract}")
    else:
        print("   ⚠️  No General Abstract found - creating new structure")
    
    print("\n3️⃣ Importing Abstract sheets...")
    for sheet in reader.abstract_sheets:
        print(f"   ✅ Imported: {sheet}")
        print(f"      📊 Setting up cost calculation formulas")
    
    print("\n4️⃣ Importing Measurement sheets...")
    for sheet in reader.measurement_sheets:
        print(f"   ✅ Imported: {sheet}")
        print(f"      📏 Setting up quantity calculation formulas")
    
    print("\n5️⃣ Rebuilding formulas and linkages...")
    pairs = reader.find_sheet_pairs()
    for abstract, measurement in pairs:
        part_name = reader.extract_part_name(abstract, "abstract")
        print(f"   🔗 Linking: {measurement} → {abstract}")
        print(f"      📊 Quantities from measurements feed into abstract amounts")
    
    print(f"   🔗 Linking Abstract totals to General Abstract")
    print(f"      📊 All part totals sum into master summary")
    
    print("\n6️⃣ Setting up protection and validation...")
    print("   🔒 Formula cells protected")
    print("   ✏️  Data entry cells unlocked")
    print("   ✅ Validation rules applied")
    
    print(f"\n✅ IMPORT COMPLETE!")
    print(f"📊 Ready for data entry and calculations")
    print(f"🔄 Real-time updates enabled")

if __name__ == "__main__":
    main()